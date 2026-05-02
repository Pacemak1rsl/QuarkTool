import sys
import re
import time
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Set

from PySide6.QtCore import QFileInfo, QPointF, QRectF, QSize, Qt, QThread, Signal, QUrl
from PySide6.QtGui import QColor, QFont, QFontDatabase, QIcon, QPainter, QPainterPath, QPen, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QAbstractSpinBox,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFormLayout,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QProgressBar,
    QPushButton,
    QFileDialog,
    QFileIconProvider,
    QSizePolicy,
    QSpinBox,
    QSplitter,
    QStackedWidget,
    QStyle,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

try:
    from PySide6.QtWebEngineWidgets import QWebEngineView

    WEB_ENGINE_AVAILABLE = True
except Exception:
    QWebEngineView = None  # type: ignore
    WEB_ENGINE_AVAILABLE = False

from .quark_api import QuarkClient
from .session_store import Account, SessionStore
from .tasks import BatchJob, TaskWorker


_ICON_CACHE: Dict[str, QIcon] = {}


def preferred_ui_families() -> List[str]:
    preferred = [
        "SF Pro Display",
        "SF Pro Text",
        "PingFang SC",
        "PingFang HK",
        "MiSans",
        "HarmonyOS Sans SC",
        "Noto Sans SC",
        "Microsoft YaHei UI",
        "Microsoft YaHei",
        "Helvetica Neue",
        "Segoe UI Variable Text",
        "Segoe UI",
    ]
    try:
        installed = set(QFontDatabase.families())
        available = [family for family in preferred if family in installed]
    except Exception:
        available = []
    return available or ["Segoe UI", "Microsoft YaHei UI"]


def _asset_path(name: str) -> Optional[Path]:
    candidates = []
    bundle_root = getattr(sys, "_MEIPASS", None)
    if bundle_root:
        candidates.append(Path(bundle_root) / "assets" / name)
    candidates.append(Path(__file__).resolve().parents[2] / "assets" / name)
    candidates.append(Path.cwd() / "assets" / name)
    for path in candidates:
        if path.exists():
            return path
    return None


def app_icon() -> QIcon:
    preferred = _asset_path("app_icon.ico") or _asset_path("pantools_icon.ico")
    if preferred:
        return QIcon(str(preferred))
    icon = QIcon()
    for size in (16, 24, 32, 48, 64, 128, 256):
        icon.addPixmap(_app_icon_pixmap(size))
    return icon


def _app_icon_pixmap(size: int) -> QPixmap:
    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
    scale = size / 64.0

    def v(value: float) -> float:
        return value * scale

    painter.setPen(Qt.PenStyle.NoPen)
    painter.setBrush(QColor("#0A84FF"))
    painter.drawRoundedRect(QRectF(v(5), v(5), v(54), v(54)), v(14), v(14))

    painter.setBrush(QColor("#FFFFFF"))
    body = QPainterPath()
    body.moveTo(v(16), v(24))
    body.lineTo(v(27), v(24))
    body.lineTo(v(31), v(29))
    body.lineTo(v(48), v(29))
    body.lineTo(v(48), v(45))
    body.lineTo(v(16), v(45))
    body.closeSubpath()
    painter.drawPath(body)

    painter.setBrush(QColor("#D9ECFF"))
    painter.drawRoundedRect(QRectF(v(20), v(33), v(24), v(6)), v(3), v(3))

    pen = QPen(QColor("#0A84FF"), max(1, int(v(3))))
    pen.setCapStyle(Qt.PenCapStyle.RoundCap)
    painter.setPen(pen)
    painter.drawLine(QPointF(v(27), v(36)), QPointF(v(37), v(36)))
    painter.end()
    return pixmap


def sidebar_brand_pixmap(size: int) -> QPixmap:
    for candidate in [Path(r"D:\mCloudDownload\PanTools.exe")]:
        try:
            if not candidate.exists():
                continue
            icon = QFileIconProvider().icon(QFileInfo(str(candidate)))
            if icon.isNull():
                continue
            pixmap = icon.pixmap(QSize(size, size))
            if not pixmap.isNull():
                return pixmap
        except Exception:
            continue
    return app_icon().pixmap(QSize(size, size))


def line_icon(name: str, color: str = "#2C2C2E", accent: str = "#007AFF", size: int = 24) -> QIcon:
    key = "%s:%s:%s:%d" % (name, color, accent, size)
    cached = _ICON_CACHE.get(key)
    if cached:
        return cached

    pixmap = QPixmap(size, size)
    pixmap.fill(Qt.GlobalColor.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)

    scale = size / 24.0

    def n(value: float) -> float:
        return value * scale

    def point(x: float, y: float) -> QPointF:
        return QPointF(n(x), n(y))

    def line(x1: float, y1: float, x2: float, y2: float) -> None:
        painter.drawLine(point(x1, y1), point(x2, y2))

    def rect(x: float, y: float, w: float, h: float, radius: float = 2.5) -> None:
        painter.drawRoundedRect(QRectF(n(x), n(y), n(w), n(h)), n(radius), n(radius))

    def ellipse(x: float, y: float, w: float, h: float) -> None:
        painter.drawEllipse(QRectF(n(x), n(y), n(w), n(h)))

    def path(points: List[tuple], close: bool = False) -> None:
        if not points:
            return
        painter_path = QPainterPath(point(points[0][0], points[0][1]))
        for x, y in points[1:]:
            painter_path.lineTo(point(x, y))
        if close:
            painter_path.closeSubpath()
        painter.drawPath(painter_path)

    pen = QPen(QColor(color), max(1, int(round(size / 12.0))))
    pen.setCapStyle(Qt.PenCapStyle.RoundCap)
    pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
    painter.setPen(pen)
    painter.setBrush(Qt.BrushStyle.NoBrush)

    if name == "app":
        painter.end()
        icon = app_icon()
        _ICON_CACHE[key] = icon
        return icon
    if name == "folder_solid":
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor("#FFCD38"))
        body = QPainterPath(point(3.3, 8.2))
        body.lineTo(point(8.6, 8.2))
        body.lineTo(point(10.7, 10.1))
        body.lineTo(point(20.5, 10.1))
        body.lineTo(point(20.5, 18.5))
        body.lineTo(point(3.3, 18.5))
        body.closeSubpath()
        painter.drawPath(body)
        painter.setBrush(QColor("#FFD85A"))
        tab = QPainterPath(point(3.8, 7.1))
        tab.lineTo(point(8.5, 7.1))
        tab.lineTo(point(10.2, 8.8))
        tab.lineTo(point(15.1, 8.8))
        tab.lineTo(point(15.1, 10.2))
        tab.lineTo(point(3.8, 10.2))
        tab.closeSubpath()
        painter.drawPath(tab)
        accent_pen = QPen(QColor("#E2A900"), max(1, int(round(size / 18.0))))
        accent_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        accent_pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        painter.setPen(accent_pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        path([(3.6, 8.2), (8.4, 8.2), (10.5, 10.1), (20.2, 10.1), (20.2, 18.2), (3.6, 18.2)], True)
    elif name == "folder":
        path([(3.5, 7.5), (8.5, 7.5), (10.5, 9.5), (20.5, 9.5), (20.5, 18.5), (3.5, 18.5)], True)
    elif name == "file":
        path([(6.5, 4.5), (13.5, 4.5), (17.5, 8.5), (17.5, 19.5), (6.5, 19.5)], True)
        line(13.5, 4.8, 13.5, 8.5)
        line(13.5, 8.5, 17.1, 8.5)
    elif name == "share":
        line(8.5, 12, 15.8, 7.8)
        line(8.5, 12, 15.8, 16.2)
        ellipse(4.4, 9.8, 4.4, 4.4)
        ellipse(15.4, 4.7, 4.4, 4.4)
        ellipse(15.4, 14.9, 4.4, 4.4)
    elif name == "transfer":
        rect(4.5, 15.5, 15, 4)
        line(12, 4.5, 12, 13)
        path([(8.5, 9.8), (12, 13.3), (15.5, 9.8)])
    elif name == "rename":
        rect(5.5, 5, 9.5, 14)
        line(8, 9, 12.5, 9)
        line(8, 12, 11, 12)
        line(14.2, 17.8, 19.5, 12.5)
        line(17.6, 10.6, 19.5, 12.5)
    elif name == "copy":
        rect(7.5, 5, 10.5, 12.5)
        rect(4.5, 8, 10.5, 12.5)
    elif name == "move":
        line(5, 12, 19, 12)
        path([(15.5, 8.5), (19, 12), (15.5, 15.5)])
        line(12, 5, 12, 19)
        path([(8.5, 8.5), (12, 5), (15.5, 8.5)])
        path([(8.5, 15.5), (12, 19), (15.5, 15.5)])
    elif name == "myshare":
        ellipse(8.4, 4.5, 7.2, 7.2)
        path([(5, 19), (6.6, 15.3), (12, 14), (17.4, 15.3), (19, 19)])
    elif name == "link":
        rect(3.8, 8, 8.6, 6.2, 3)
        rect(11.6, 9.8, 8.6, 6.2, 3)
        line(9.8, 12.1, 14.2, 12.1)
    elif name == "settings":
        line(4, 7, 20, 7)
        line(4, 12, 20, 12)
        line(4, 17, 20, 17)
        ellipse(8, 5.4, 3.2, 3.2)
        ellipse(14.8, 10.4, 3.2, 3.2)
        ellipse(6, 15.4, 3.2, 3.2)
    elif name == "back":
        path([(14.5, 6), (8.5, 12), (14.5, 18)])
    elif name == "refresh":
        painter.drawArc(QRectF(n(5), n(5), n(14), n(14)), 35 * 16, 285 * 16)
        path([(16.6, 4.7), (19.1, 5.2), (18.3, 7.8)])
    elif name == "home":
        path([(4.8, 11.5), (12, 5), (19.2, 11.5)])
        path([(7.2, 10), (7.2, 19), (16.8, 19), (16.8, 10)])
    elif name == "folder_plus":
        path([(3.5, 7.5), (8.5, 7.5), (10.5, 9.5), (20.5, 9.5), (20.5, 18.5), (3.5, 18.5)], True)
        line(12, 12, 12, 16)
        line(10, 14, 14, 14)
    elif name == "account":
        ellipse(7.2, 5, 6, 6)
        path([(4.5, 18.5), (5.8, 14.8), (10.2, 13.6), (14.2, 14.8)])
        line(17.5, 9.5, 17.5, 15.5)
        line(14.5, 12.5, 20.5, 12.5)
    elif name == "search":
        ellipse(5, 5, 10, 10)
        line(13.4, 13.4, 19, 19)
    elif name == "check":
        path([(5, 12.5), (9.2, 16.5), (18.8, 7.5)])
    elif name == "trash":
        line(6, 8, 18, 8)
        line(9, 5.5, 15, 5.5)
        rect(7.5, 8, 9, 11)
        line(10.5, 11, 10.5, 16)
        line(13.5, 11, 13.5, 16)
    elif name == "plus":
        line(12, 6, 12, 18)
        line(6, 12, 18, 12)
    else:
        ellipse(6, 6, 12, 12)

    if accent and name in {"share", "transfer", "folder_plus", "account", "search", "check", "plus"}:
        accent_pen = QPen(QColor(accent), max(1, int(round(size / 14.0))))
        accent_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        painter.setPen(accent_pen)
        if name == "share":
            ellipse(15.8, 5.1, 3.6, 3.6)
        elif name == "folder_plus":
            line(12, 12.2, 12, 15.8)
            line(10.2, 14, 13.8, 14)

    painter.end()
    icon = QIcon(pixmap)
    _ICON_CACHE[key] = icon
    return icon


class ApiThread(QThread):
    success = Signal(object)
    error = Signal(str)

    def __init__(self, func: Callable[..., Any], *args: Any, **kwargs: Any) -> None:
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs

    def run(self) -> None:
        try:
            self.success.emit(self.func(*self.args, **self.kwargs))
        except Exception as exc:
            self.error.emit(str(exc))


class LoginDialog(QDialog):
    def __init__(self, store: SessionStore, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.store = store
        self.cookies: Dict[str, str] = {}
        self.setWindowTitle("扫码登录夸克网盘")
        self.resize(980, 720)
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        header = QLabel("打开页面后用手机夸克扫码登录，登录完成再点“保存登录态”。")
        header.setObjectName("mutedLabel")
        layout.addWidget(header)

        if WEB_ENGINE_AVAILABLE:
            self.web = QWebEngineView(self)  # type: ignore
            self.web.setUrl(QUrl("https://pan.quark.cn/"))
            layout.addWidget(self.web, 1)
            profile = self.web.page().profile()
            cookie_store = profile.cookieStore()
            cookie_store.cookieAdded.connect(self._on_cookie_added)
            try:
                cookie_store.loadAllCookies()
            except Exception:
                pass
        else:
            message = QLabel("当前环境没有 Qt WebEngine，先用“手动粘贴 Cookie”保存登录态。")
            message.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.addWidget(message, 1)

        buttons = QHBoxLayout()
        paste_btn = QPushButton("手动粘贴 Cookie")
        save_btn = QPushButton("保存登录态")
        save_btn.setObjectName("primaryButton")
        buttons.addStretch(1)
        buttons.addWidget(paste_btn)
        buttons.addWidget(save_btn)
        layout.addLayout(buttons)

        paste_btn.clicked.connect(self._paste_cookie)
        save_btn.clicked.connect(self._save_web_cookie)

    def _on_cookie_added(self, cookie: Any) -> None:
        try:
            domain = cookie.domain()
            if "quark.cn" not in domain and "uc.cn" not in domain:
                return
            name = bytes(cookie.name()).decode("utf-8", "ignore")
            value = bytes(cookie.value()).decode("utf-8", "ignore")
            if name and value:
                self.cookies[name] = value
        except Exception:
            return

    def _save_web_cookie(self) -> None:
        cookie_header = self._cookie_header()
        if len(cookie_header) < 20:
            QMessageBox.information(self, "还没有登录态", "还没有捕获到足够的 Cookie。可以扫码登录后再试，或手动粘贴。")
            return
        self._save_cookie(cookie_header)

    def _paste_cookie(self) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("手动粘贴 Cookie")
        dialog.resize(720, 360)
        layout = QVBoxLayout(dialog)
        edit = QPlainTextEdit()
        edit.setPlaceholderText("粘贴完整 Cookie，例如：k1=v1; k2=v2")
        layout.addWidget(edit, 1)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        cookie_header = edit.toPlainText().strip()
        if len(cookie_header) < 20:
            QMessageBox.warning(self, "Cookie 太短", "请输入完整 Cookie。")
            return
        self._save_cookie(cookie_header)

    def _save_cookie(self, cookie_header: str) -> None:
        default_name = "夸克账号 %d" % (len(self.store.list_accounts()) + 1)
        name, ok = QInputDialog.getText(self, "账号名称", "给这个登录态取个名字：", text=default_name)
        if not ok:
            return
        self.store.add_or_update_account(name.strip() or default_name, cookie_header)
        self.accept()

    def _cookie_header(self) -> str:
        return "; ".join("%s=%s" % (name, value) for name, value in sorted(self.cookies.items()))


class ShareDialog(QDialog):
    def __init__(self, count: int, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("批量分享")
        layout = QFormLayout(self)
        self.password = QLineEdit()
        self.password.setPlaceholderText("可留空")
        self.delay_min = 5
        self.delay_max = QSpinBox()
        self.delay_max.setRange(6, 3600)
        self.delay_max.setValue(10)
        layout.addRow(QLabel("已选择 %d 个文件/文件夹" % count))
        layout.addRow(QLabel("有效期：永久有效"))
        layout.addRow("提取密码", self.password)
        layout.addRow("延时秒 (5 ~)", self.delay_max)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addRow(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def delay_range(self) -> tuple:
        return self.delay_min, max(self.delay_max.value(), self.delay_min)


class BulkTransferDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("批量转存")
        self.resize(760, 560)
        layout = QVBoxLayout(self)
        self.links = QPlainTextEdit()
        self.links.setPlaceholderText("每行一个夸克分享链接，可附带提取码，例如：\nhttps://pan.quark.cn/s/xxxx 提取码: abcd")
        layout.addWidget(self.links, 1)

        form = QFormLayout()
        self.target_folder = QLineEdit("root")
        self.auto_share = QCheckBox("转存后尝试直接创建分享链接")
        self.delay_min = 5
        self.delay_max = QSpinBox()
        self.delay_max.setRange(6, 3600)
        self.delay_max.setValue(10)
        form.addRow("保存到文件夹 FID", self.target_folder)
        form.addRow("", self.auto_share)
        form.addRow(QLabel("自动分享有效期：永久有效"))
        form.addRow("延时秒 (5 ~)", self.delay_max)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def jobs(self) -> List[BatchJob]:
        result: List[BatchJob] = []
        for raw_line in self.links.toPlainText().splitlines():
            line = raw_line.strip()
            if not line:
                continue
            match = re.search(r"https?://pan\.quark\.cn/s/[a-zA-Z0-9]+", line)
            if not match:
                continue
            code_match = re.search(r"(?:提取码|密码|passcode|code)[:：\s]*([a-zA-Z0-9]+)", line)
            passcode = code_match.group(1) if code_match else ""
            kind = "transfer_share" if self.auto_share.isChecked() else "transfer"
            result.append(
                BatchJob(
                    kind=kind,
                    label="转存 %s" % match.group(0),
                    payload={
                        "share_url": match.group(0),
                        "passcode": passcode,
                        "target_folder_id": self.target_folder.text().strip() or "root",
                        "expire_days": 0,
                    },
                )
            )
        return result

    def delay_range(self) -> tuple:
        return self.delay_min, max(self.delay_max.value(), self.delay_min)


class CopyDialog(QDialog):
    def __init__(self, source_fid: str = "", parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("批量复制到多个文件夹")
        self.resize(680, 460)
        layout = QFormLayout(self)
        self.source_fid = QLineEdit(source_fid)
        self.targets = QPlainTextEdit()
        self.targets.setPlaceholderText("每行一个目标文件夹 FID；根目录可填 root")
        self.delay_min = 5
        self.delay_max = QSpinBox()
        self.delay_max.setRange(6, 3600)
        self.delay_max.setValue(10)
        layout.addRow("源文件 FID", self.source_fid)
        layout.addRow("目标文件夹 FID", self.targets)
        layout.addRow("延时秒 (5 ~)", self.delay_max)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addRow(buttons)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def jobs(self) -> List[BatchJob]:
        fid = self.source_fid.text().strip()
        jobs = []
        for target in self.targets.toPlainText().splitlines():
            target = target.strip()
            if target and fid:
                jobs.append(BatchJob("copy", "复制到 %s" % target, {"fid": fid, "target_folder_id": target}))
        return jobs

    def delay_range(self) -> tuple:
        return self.delay_min, max(self.delay_max.value(), self.delay_min)


class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Pan Tools")
        self.setWindowIcon(app_icon())
        self.store = SessionStore()
        self.client: Optional[QuarkClient] = None
        self.accounts: List[Account] = []
        self.files: List[Dict[str, Any]] = []
        self.current_folder = "root"
        self.folder_history: List[str] = []
        self.copy_sources: List[Dict[str, Any]] = []
        self.copy_targets: List[Dict[str, Any]] = []
        self.copy_target_status: Dict[str, str] = {}
        self.copy_target_progress: Dict[str, Dict[str, int]] = {}
        self.share_items: List[Dict[str, Any]] = []
        self.share_results: Dict[str, Dict[str, Any]] = {}
        self.transfer_results: List[Dict[str, Any]] = []
        self.rename_items: List[Dict[str, Any]] = []
        self.rename_status: Dict[str, str] = {}
        self.move_items: List[Dict[str, Any]] = []
        self.move_target: Optional[Dict[str, Any]] = None
        self.move_status: Dict[str, str] = {}
        self.recycle_items: List[Dict[str, Any]] = []
        self.recycle_visible_items: List[Dict[str, Any]] = []
        self.recycle_status: Dict[str, str] = {}
        self.account_health: Dict[str, Dict[str, str]] = {}
        self.nav_buttons: Dict[str, QPushButton] = {}
        self.current_nav = ""
        self.page_size = 60
        self.file_page = 0
        self.checked_file_ids: Set[str] = set()
        self.table_mode = "files"
        self.worker: Optional[TaskWorker] = None
        self.last_job_kinds: Set[str] = set()
        self.api_thread: Optional[ApiThread] = None
        self._build_ui()
        self._apply_style()
        self._update_file_footer()
        self._reload_accounts()

    def _build_ui(self) -> None:
        root = QWidget()
        root.setObjectName("appRoot")
        root_layout = QHBoxLayout(root)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)
        root_layout.addWidget(self._sidebar())
        root_layout.addWidget(self._main_area(), 1)
        self.setCentralWidget(root)

    def _sidebar(self) -> QWidget:
        sidebar = QFrame()
        sidebar.setObjectName("sidebar")
        sidebar.setFixedWidth(180)
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        brand_row = QWidget()
        header_height = 61
        brand_row.setObjectName("sidebarHeader")
        brand_row.setFixedHeight(header_height)
        brand_row_layout = QHBoxLayout(brand_row)
        brand_row_layout.setContentsMargins(12, 14, 12, 10)
        brand_row_layout.setSpacing(0)
        brand_group = QWidget()
        brand_group_layout = QHBoxLayout(brand_group)
        brand_group_layout.setContentsMargins(0, 0, 0, 0)
        brand_group_layout.setSpacing(6)
        brand_icon = QLabel()
        brand_icon.setObjectName("brandIcon")
        brand_icon.setPixmap(sidebar_brand_pixmap(35))
        brand_icon.setFixedSize(35, 35)
        brand = QLabel("PanTools")
        brand.setObjectName("brand")
        brand_group_layout.addWidget(brand_icon)
        brand_group_layout.addWidget(brand)
        brand_row_layout.addWidget(brand_group)
        brand_row_layout.addStretch(1)
        layout.addWidget(brand_row)

        nav_container = QWidget()
        nav_container.setObjectName("sidebarNav")
        nav_layout = QVBoxLayout(nav_container)
        nav_layout.setContentsMargins(8, 12, 8, 18)
        nav_layout.setSpacing(8)
        items = [
            ("files", "文件管理", self.refresh_files, "folder_solid", "#F5B400", "#E2A900"),
            ("share", "批量分享", self.show_share_workspace, "share", "#0A84FF", "#5AC8FA"),
            ("transfer", "批量转存", self.show_transfer_workspace, "transfer", "#34C759", "#30D158"),
            ("rename", "批量重命名", self.show_rename_workspace, "rename", "#AF52DE", "#BF5AF2"),
            ("copy", "批量复制", self.show_copy_workspace, "copy", "#5AC8FA", "#64D2FF"),
            ("move", "批量移动", self.show_move_workspace, "move", "#FF9F0A", "#FFB340"),
            ("trash", "回收站", self.show_trash_workspace, "trash", "#FF6B5A", "#FF3B30"),
            ("settings", "账号池", self.show_settings, "settings", "#30B0C7", "#64D2FF"),
        ]
        for key, text, slot, icon_name, icon_color, icon_accent in items:
            button = QPushButton(text)
            button.setObjectName("navButton")
            button.setIcon(line_icon(icon_name, icon_color, icon_accent))
            button.setIconSize(QSize(20, 20))
            button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            button.setProperty("active", False)
            button.clicked.connect(slot)
            self.nav_buttons[key] = button
            nav_layout.addWidget(button)
        self._set_active_nav("files")

        nav_layout.addStretch(1)
        version = QLabel("v1.0")
        version.setObjectName("version")
        nav_layout.addWidget(version)
        layout.addWidget(nav_container, 1)
        return sidebar

    def _main_area(self) -> QWidget:
        main = QWidget()
        layout = QVBoxLayout(main)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(self._topbar())
        layout.addWidget(self._toolbar())

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setObjectName("mainSplitter")
        self.content_stack = QStackedWidget()
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["", "文件名", "FID", "大小", "修改时间", "操作"])
        self.table.verticalHeader().setVisible(False)
        self._polish_table(self.table, 44)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 48)
        self.table.setColumnWidth(2, 300)
        self.table.setColumnWidth(5, 180)
        self.table.cellDoubleClicked.connect(self.open_row)
        self.table.itemChanged.connect(self._handle_table_item_changed)
        self.file_list_page = QWidget()
        content_layout = QVBoxLayout(self.file_list_page)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        content_layout.addWidget(self.table, 1)
        self.file_footer = QFrame()
        self.file_footer.setObjectName("fileFooter")
        footer_layout = QHBoxLayout(self.file_footer)
        footer_layout.setContentsMargins(18, 10, 18, 10)
        footer_layout.setSpacing(8)
        self.select_page_folders_btn = QPushButton("全选")
        self.select_page_folders_btn.setObjectName("smallGrayButton")
        self.select_page_folders_btn.clicked.connect(self.toggle_select_current_page)
        self.page_info_label = QLabel("第 1 / 1 页")
        self.page_info_label.setObjectName("quota")
        self.page_summary_label = QLabel("0 / 0")
        self.page_summary_label.setObjectName("quota")
        self.page_prev_btn = QPushButton("上页")
        self.page_prev_btn.setObjectName("smallGrayButton")
        self.page_prev_btn.clicked.connect(self.goto_prev_page)
        self.page_next_btn = QPushButton("下页")
        self.page_next_btn.setObjectName("smallGrayButton")
        self.page_next_btn.clicked.connect(self.goto_next_page)
        footer_layout.addWidget(self.select_page_folders_btn)
        footer_layout.addStretch(1)
        footer_layout.addWidget(self.page_summary_label)
        footer_layout.addWidget(self.page_info_label)
        footer_layout.addWidget(self.page_prev_btn)
        footer_layout.addWidget(self.page_next_btn)
        content_layout.addWidget(self.file_footer)
        self.content_stack.addWidget(self.file_list_page)
        self.copy_page = self._copy_workspace()
        self.content_stack.addWidget(self.copy_page)
        self.share_page = self._share_workspace()
        self.content_stack.addWidget(self.share_page)
        self.transfer_page = self._transfer_workspace()
        self.content_stack.addWidget(self.transfer_page)
        self.rename_page = self._rename_workspace()
        self.content_stack.addWidget(self.rename_page)
        self.move_page = self._move_workspace()
        self.content_stack.addWidget(self.move_page)
        self.trash_page = self._trash_workspace()
        self.content_stack.addWidget(self.trash_page)
        self.settings_page = self._settings_workspace()
        self.content_stack.addWidget(self.settings_page)

        bottom = QWidget()
        bottom.setObjectName("statusPanel")
        bottom_layout = QVBoxLayout(bottom)
        bottom_layout.setContentsMargins(18, 12, 18, 16)
        bottom_layout.setSpacing(10)
        control_row = QHBoxLayout()
        self.hint = QLabel("提示：勾选文件后可加入批量分享、复制、移动任务。批量任务默认每步随机等待 5-10 秒。")
        self.progress = QProgressBar()
        self.progress.setFixedWidth(260)
        self.stop_btn = QPushButton("停止")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_worker)
        control_row.addWidget(self.hint, 1)
        control_row.addWidget(self.progress)
        control_row.addWidget(self.stop_btn)
        self.log_box = QTextEdit()
        self.log_box.setObjectName("logBox")
        self.log_box.setReadOnly(True)
        self.log_box.setFixedHeight(140)
        bottom_layout.addLayout(control_row)
        bottom_layout.addWidget(self.log_box)

        splitter.addWidget(self.content_stack)
        splitter.addWidget(bottom)
        splitter.setSizes([620, 190])
        layout.addWidget(splitter, 1)
        return main

    def _compact_delay_spinbox(self, value: int) -> QSpinBox:
        spin = QSpinBox()
        spin.setObjectName("compactSpin")
        spin.setRange(6, 3600)
        spin.setValue(value)
        spin.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        spin.setAlignment(Qt.AlignmentFlag.AlignCenter)
        spin.setFixedSize(34, 34)
        return spin

    def _compact_label(self, text: str, width: int = 10) -> QLabel:
        label = QLabel(text)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setFixedWidth(width)
        return label

    def _copy_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(4)
        title = QLabel("批量复制")
        title.setObjectName("pageTitle")
        self.copy_delay_min = 5
        self.copy_delay_max = self._compact_delay_spinbox(10)
        run_btn = QPushButton("开始复制")
        run_btn.setObjectName("successButton")
        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("dangerButton")
        run_btn.clicked.connect(self.run_copy_workspace)
        clear_btn.clicked.connect(self.clear_copy_workspace)
        sec = self._compact_label("秒", 18)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(QLabel("延时 5 -"))
        header.addWidget(self.copy_delay_max)
        header.addWidget(sec)
        header.addWidget(run_btn)
        header.addWidget(clear_btn)
        layout.addLayout(header)
        self.copy_summary = QLabel("先选中多个源文件或文件夹，再选择多个目标文件夹。")
        self.copy_summary.setObjectName("sectionCaption")
        layout.addWidget(self.copy_summary)

        tables = QSplitter(Qt.Orientation.Horizontal)
        self.copy_source_table = QTableWidget(0, 2)
        self.copy_source_table.setHorizontalHeaderLabels(["文件名[被复制]", "操作"])
        self.copy_target_table = QTableWidget(0, 3)
        self.copy_target_table.setHorizontalHeaderLabels(["文件名[复制到]", "状态", "操作"])
        for table in [self.copy_source_table, self.copy_target_table]:
            table.verticalHeader().setVisible(False)
            self._polish_table(table, 44)
            table.setAlternatingRowColors(True)
            table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
            table.horizontalHeader().setStretchLastSection(False)
        self.copy_source_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.copy_source_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.copy_source_table.setColumnWidth(1, 110)
        self.copy_target_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.copy_target_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.copy_target_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.copy_target_table.setColumnWidth(1, 100)
        self.copy_target_table.setColumnWidth(2, 110)
        tables.addWidget(self.copy_source_table)
        tables.addWidget(self.copy_target_table)
        tables.setSizes([430, 920])
        layout.addWidget(tables, 1)
        return page

    def _share_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(4)
        title = QLabel("批量分享")
        title.setObjectName("pageTitle")
        self.share_delay_min = 5
        self.share_delay_max = self._compact_delay_spinbox(10)
        run_btn = QPushButton("开始分享")
        run_btn.setObjectName("successButton")
        run_btn.setMinimumWidth(96)
        export_btn = QPushButton("导出 Excel")
        export_btn.setObjectName("primaryButton")
        export_btn.setMinimumWidth(106)
        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("dangerButton")
        clear_btn.setMinimumWidth(80)
        run_btn.clicked.connect(self.run_share_workspace)
        export_btn.clicked.connect(self.export_share_results)
        clear_btn.clicked.connect(self.clear_share_workspace)
        sec = self._compact_label("秒", 18)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(QLabel("延时 5 -"))
        header.addWidget(self.share_delay_max)
        header.addWidget(sec)
        header.addWidget(run_btn)
        header.addWidget(export_btn)
        header.addWidget(clear_btn)
        layout.addLayout(header)
        self.share_summary = QLabel("批量分享会默认创建永久有效链接。")
        self.share_summary.setObjectName("sectionCaption")
        layout.addWidget(self.share_summary)

        self.share_table = QTableWidget(0, 5)
        self.share_table.setHorizontalHeaderLabels(["文件名", "状态", "分享链接", "提取码", "操作"])
        self.share_table.verticalHeader().setVisible(False)
        self._polish_table(self.share_table, 44)
        self.share_table.setAlternatingRowColors(True)
        self.share_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.share_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.share_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.share_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.share_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.share_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.share_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.share_table.setColumnWidth(1, 100)
        self.share_table.setColumnWidth(3, 90)
        self.share_table.setColumnWidth(4, 110)
        layout.addWidget(self.share_table, 1)
        return page

    def _transfer_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header_layout = QHBoxLayout()
        header_layout.setSpacing(10)
        title = QLabel("批量转存")
        title.setObjectName("pageTitle")
        header_layout.addWidget(title)
        header_layout.addStretch(1)
        view_template_btn = QPushButton("查看导入Excel模板")
        view_template_btn.setObjectName("primaryButton")
        view_template_btn.clicked.connect(self.view_import_template)
        header_layout.addWidget(view_template_btn)
        layout.addLayout(header_layout)

        self.transfer_summary = QLabel("支持粘贴链接或导入 Excel，适合批量转存并分享。")
        self.transfer_summary.setObjectName("sectionCaption")
        layout.addWidget(self.transfer_summary)

        self.transfer_links = QPlainTextEdit()
        self.transfer_links.setPlaceholderText(
            "每行一个夸克分享链接，可带提取码：\n"
            "https://pan.quark.cn/s/xxxx 提取码: abcd"
        )
        layout.addWidget(self.transfer_links, 2)

        options = QHBoxLayout()
        options.setSpacing(4)
        self.transfer_target_folder = QLineEdit("971ba8e22a4f4c7b820ee15a11da1465")
        self.transfer_auto_share = QCheckBox("转存后自动分享")
        self.transfer_delay_min = 5
        self.transfer_delay_max = self._compact_delay_spinbox(6)
        run_btn = QPushButton("开始转存")
        run_btn.setObjectName("successButton")
        import_btn = QPushButton("导入 Excel")
        import_btn.setObjectName("primaryButton")
        export_btn = QPushButton("导出 Excel")
        export_btn.setObjectName("primaryButton")
        clear_btn = QPushButton("清空结果")
        clear_btn.setObjectName("dangerButton")
        run_btn.clicked.connect(self.run_transfer_workspace)
        import_btn.clicked.connect(self.import_transfer_excel)
        export_btn.clicked.connect(self.export_transfer_results)
        clear_btn.clicked.connect(self.clear_transfer_results)
        sec = self._compact_label("秒", 18)

        options.addWidget(QLabel("保存到文件夹 FID"))
        options.addWidget(self.transfer_target_folder, 2)
        options.addWidget(self.transfer_auto_share)
        options.addWidget(QLabel("延时 5 -"))
        options.addWidget(self.transfer_delay_max)
        options.addWidget(sec)
        options.addWidget(import_btn)
        options.addWidget(run_btn)
        options.addWidget(export_btn)
        options.addWidget(clear_btn)
        layout.addLayout(options)

        self.transfer_table = QTableWidget(0, 5)
        self.transfer_table.setHorizontalHeaderLabels(["文件名称", "原夸克链接", "状态", "新分享链接", "错误信息"])
        self.transfer_table.verticalHeader().setVisible(False)
        self._polish_table(self.transfer_table, 44)
        self.transfer_table.setAlternatingRowColors(True)
        self.transfer_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.transfer_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        transfer_header = self.transfer_table.horizontalHeader()
        transfer_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        transfer_header.setStretchLastSection(False)
        transfer_header.setMinimumSectionSize(64)
        self.transfer_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.transfer_table.setColumnWidth(0, 210)
        self.transfer_table.setColumnWidth(1, 280)
        self.transfer_table.setColumnWidth(2, 92)
        self.transfer_table.setColumnWidth(3, 280)
        self.transfer_table.setColumnWidth(4, 260)
        layout.addWidget(self.transfer_table, 2)
        return page

    def _rename_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(4)
        title = QLabel("批量重命名")
        title.setObjectName("pageTitle")
        self.rename_delay_min = 5
        self.rename_delay_max = self._compact_delay_spinbox(10)
        run_btn = QPushButton("开始重命名")
        run_btn.setObjectName("successButton")
        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("dangerButton")
        run_btn.clicked.connect(self.run_rename_workspace)
        clear_btn.clicked.connect(self.clear_rename_workspace)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(QLabel("延时 5 -"))
        header.addWidget(self.rename_delay_max)
        header.addWidget(self._compact_label("秒", 18))
        header.addWidget(run_btn)
        header.addWidget(clear_btn)
        layout.addLayout(header)
        self.rename_summary = QLabel("双击“新文件名”列即可直接编辑。")
        self.rename_summary.setObjectName("sectionCaption")
        layout.addWidget(self.rename_summary)

        self.rename_table = QTableWidget(0, 4)
        self.rename_table.setHorizontalHeaderLabels(["原文件名", "新文件名", "状态", "操作"])
        self.rename_table.verticalHeader().setVisible(False)
        self._polish_table(self.rename_table, 44)
        self.rename_table.setAlternatingRowColors(True)
        self.rename_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        self.rename_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.rename_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.rename_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.rename_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.rename_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.rename_table.setColumnWidth(2, 100)
        self.rename_table.setColumnWidth(3, 110)
        layout.addWidget(self.rename_table, 1)
        return page

    def _move_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(4)
        title = QLabel("批量移动")
        title.setObjectName("pageTitle")
        self.move_target_fid = QLineEdit()
        self.move_target_fid.setPlaceholderText("目标文件夹 FID")
        self.move_target_fid.setMinimumWidth(260)
        self.move_target_fid.textChanged.connect(lambda _: self.render_move_workspace())
        self.move_target_name = QLabel("未设置目标")
        self.move_target_name.setObjectName("quota")
        self.move_delay_min = 5
        self.move_delay_max = self._compact_delay_spinbox(10)
        run_btn = QPushButton("开始移动")
        run_btn.setObjectName("successButton")
        clear_btn = QPushButton("清空")
        clear_btn.setObjectName("dangerButton")
        run_btn.clicked.connect(self.run_move_workspace)
        clear_btn.clicked.connect(self.clear_move_workspace)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(QLabel("目标FID"))
        header.addWidget(self.move_target_fid, 1)
        header.addWidget(self.move_target_name)
        header.addWidget(QLabel("延时 5 -"))
        header.addWidget(self.move_delay_max)
        header.addWidget(self._compact_label("秒", 18))
        header.addWidget(run_btn)
        header.addWidget(clear_btn)
        layout.addLayout(header)
        self.move_summary = QLabel("可在文件管理中先选中项目，再指定目标文件夹。")
        self.move_summary.setObjectName("sectionCaption")
        layout.addWidget(self.move_summary)

        self.move_table = QTableWidget(0, 4)
        self.move_table.setHorizontalHeaderLabels(["文件名[被移动]", "目标文件夹", "状态", "操作"])
        self.move_table.verticalHeader().setVisible(False)
        self._polish_table(self.move_table, 44)
        self.move_table.setAlternatingRowColors(True)
        self.move_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.move_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.move_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.move_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.move_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.move_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.move_table.setColumnWidth(2, 100)
        self.move_table.setColumnWidth(3, 110)
        layout.addWidget(self.move_table, 1)
        return page

    def _trash_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(8)
        title = QLabel("回收站管理")
        title.setObjectName("pageTitle")
        self.trash_delay_min = 5
        self.trash_delay_max = self._compact_delay_spinbox(10)
        refresh_btn = QPushButton("刷新")
        refresh_btn.setObjectName("smallGrayButton")
        self.trash_select_btn = QPushButton("全选")
        self.trash_select_btn.setObjectName("smallGrayButton")
        self.trash_filter_combo = QComboBox()
        self.trash_filter_combo.setObjectName("filterCombo")
        self.trash_filter_combo.addItem("全部项目", "all")
        self.trash_filter_combo.addItem("1天内到期", "lt1")
        self.trash_filter_combo.addItem("3天内到期", "lt3")
        self.trash_filter_combo.addItem("7天内到期", "lt7")
        self.trash_filter_combo.addItem("7天以上", "gt7")
        self.trash_filter_combo.addItem("已过期", "expired")
        self.trash_filter_combo.setMinimumWidth(128)
        restore_btn = QPushButton("恢复选中")
        restore_btn.setObjectName("successButton")
        remove_btn = QPushButton("彻底删除")
        remove_btn.setObjectName("dangerButton")
        clear_btn = QPushButton("清空回收站")
        clear_btn.setObjectName("dangerButton")
        refresh_btn.clicked.connect(self.show_trash_workspace)
        self.trash_select_btn.clicked.connect(self.toggle_select_recycle_visible)
        self.trash_filter_combo.currentIndexChanged.connect(self.render_trash_workspace)
        restore_btn.clicked.connect(self.recover_selected_recycle)
        remove_btn.clicked.connect(self.remove_selected_recycle)
        clear_btn.clicked.connect(self.clear_recycle_bin)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(QLabel("筛选"))
        header.addWidget(self.trash_filter_combo)
        header.addWidget(QLabel("延时 5 -"))
        header.addWidget(self.trash_delay_max)
        header.addWidget(self._compact_label("秒", 18))
        header.addWidget(refresh_btn)
        header.addWidget(self.trash_select_btn)
        header.addWidget(restore_btn)
        header.addWidget(remove_btn)
        header.addWidget(clear_btn)
        layout.addLayout(header)

        self.trash_summary = QLabel("回收站共 0 项")
        self.trash_summary.setObjectName("sectionCaption")
        layout.addWidget(self.trash_summary)

        self.trash_table = QTableWidget(0, 5)
        self.trash_table.setHorizontalHeaderLabels(["文件名", "删除时间", "剩余时间", "状态", "操作"])
        self.trash_table.verticalHeader().setVisible(False)
        self._polish_table(self.trash_table, 44)
        self.trash_table.setAlternatingRowColors(True)
        self.trash_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.trash_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.trash_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.trash_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.trash_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.trash_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.trash_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.trash_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.trash_table.setColumnWidth(0, 340)
        self.trash_table.setColumnWidth(1, 200)
        self.trash_table.setColumnWidth(2, 110)
        self.trash_table.setColumnWidth(3, 90)
        self.trash_table.itemSelectionChanged.connect(self._update_trash_selection_state)
        layout.addWidget(self.trash_table, 1)
        return page

    def _settings_workspace(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(8)
        title = QLabel("账号池")
        title.setObjectName("pageTitle")
        add_btn = QPushButton("添加账号")
        add_btn.setObjectName("primaryButton")
        check_btn = QPushButton("检测账号")
        check_btn.setObjectName("smallGrayButton")
        add_btn.clicked.connect(self.open_login)
        check_btn.clicked.connect(self.refresh_account_pool_statuses)
        header.addWidget(title)
        header.addStretch(1)
        header.addWidget(check_btn)
        header.addWidget(add_btn)
        layout.addLayout(header)

        self.account_pool_summary = QLabel("已保存 0 个账号")
        self.account_pool_summary.setObjectName("sectionCaption")
        layout.addWidget(self.account_pool_summary)

        self.account_table = QTableWidget(0, 6)
        self.account_table.setHorizontalHeaderLabels(["账号名称", "当前", "状态", "创建时间", "登录态", "操作"])
        self.account_table.verticalHeader().setVisible(False)
        self._polish_table(self.account_table, 44)
        self.account_table.setAlternatingRowColors(True)
        self.account_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.account_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.account_table.horizontalHeader().setStretchLastSection(True)
        self.account_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.account_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.account_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.account_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)
        self.account_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)
        self.account_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.account_table.setColumnWidth(0, 180)
        self.account_table.setColumnWidth(1, 60)
        self.account_table.setColumnWidth(2, 78)
        self.account_table.setColumnWidth(3, 148)
        self.account_table.setColumnWidth(4, 76)
        self.account_table.setColumnWidth(5, 248)
        layout.addWidget(self.account_table, 1)
        return page

    def _topbar(self) -> QWidget:
        bar = QFrame()
        bar.setObjectName("topbar")
        header_height = 61
        bar.setFixedHeight(header_height)
        bar.setMinimumWidth(1030)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(16, 14, 16, 10)
        layout.setSpacing(8)

        self.add_account_btn = QPushButton("添加账号")
        self.add_account_btn.setObjectName("primaryButton")
        self.add_account_btn.setIcon(line_icon("account", "#FFFFFF", "#FFFFFF"))
        self.add_account_btn.setIconSize(QSize(18, 18))
        top_button_width = 128
        top_button_height = 36
        self.add_account_btn.setFixedSize(top_button_width, top_button_height)
        self.account_combo = QComboBox()
        self.account_combo.setObjectName("topCombo")
        self.account_combo.setFixedSize(142, top_button_height)
        self.account_combo.setIconSize(QSize(18, 18))
        self.folder_combo = QComboBox()
        self.folder_combo.setObjectName("topCombo")
        self.folder_combo.setFixedSize(top_button_width, top_button_height)
        self.folder_combo.setIconSize(QSize(18, 18))
        self.folder_combo.addItem(line_icon("folder", "#0A84FF", "#0A84FF"), "夸克网盘", "root")
        self.folder_combo.setEnabled(False)
        self.search_input = QLineEdit()
        self.search_input.setFixedWidth(260)
        self.search_input.setPlaceholderText("输入关键词，多个用 | 分割")
        self.search_btn = QPushButton("搜索")
        self.search_btn.setObjectName("primaryButton")
        self.search_btn.setIcon(line_icon("search", "#FFFFFF", "#FFFFFF"))
        self.search_btn.setIconSize(QSize(18, 18))
        self.search_btn.setMinimumWidth(96)
        self.multi_account = QCheckBox("多账号搜索")
        self.multi_account.setMinimumWidth(118)
        for widget in [
            self.add_account_btn,
            self.folder_combo,
            self.account_combo,
            self.search_btn,
            self.multi_account,
        ]:
            widget.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        self.search_input.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        layout.addWidget(self.add_account_btn)
        layout.addWidget(self.folder_combo)
        layout.addWidget(self.account_combo)
        layout.addWidget(self.search_input)
        layout.addWidget(self.search_btn)
        layout.addWidget(self.multi_account)

        self.add_account_btn.clicked.connect(self.open_login)
        self.account_combo.currentIndexChanged.connect(self._account_changed)
        self.search_btn.clicked.connect(self.search_files)
        self.search_input.returnPressed.connect(self.search_files)
        return bar

    def _toolbar(self) -> QWidget:
        bar = QFrame()
        bar.setObjectName("toolbar")
        bar.setMinimumWidth(980)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(16, 10, 16, 10)
        layout.setSpacing(8)

        back = self._tool_button(QStyle.StandardPixmap.SP_ArrowBack, "返回上级")
        refresh = self._tool_button(QStyle.StandardPixmap.SP_BrowserReload, "刷新")
        home = self._tool_button(QStyle.StandardPixmap.SP_DirHomeIcon, "根目录")
        new_folder = self._tool_button(QStyle.StandardPixmap.SP_FileDialogNewFolder, "新建文件夹")
        self.operation_combo = QComboBox()
        self.operation_combo.setMinimumWidth(260)
        self.operation_combo.setIconSize(QSize(18, 18))
        self.operation_combo.addItem(line_icon("settings", "#8E8E93"), "请选择操作类型", "")
        self.operation_combo.addItem(line_icon("share", "#007AFF", "#007AFF"), "加入到批量分享", "share_queue")
        self.operation_combo.addItem(line_icon("file", "#007AFF", "#007AFF"), "加入到被复制", "copy_source")
        self.operation_combo.addItem(line_icon("folder", "#007AFF", "#007AFF"), "加入到复制到", "copy_target")
        self.operation_combo.addItem(line_icon("rename", "#007AFF", "#007AFF"), "加入到批量重命名", "rename_queue")
        self.operation_combo.addItem(line_icon("move", "#007AFF", "#007AFF"), "加入到批量移动", "move_queue")
        self.operation_combo.addItem(line_icon("folder", "#007AFF", "#007AFF"), "设为移动到位置", "move_target")
        self.operation_combo.addItem(line_icon("trash", "#FF3B30", "#FF3B30"), "批量删除", "delete_items")
        self.apply_operation_btn = QPushButton("加入选中")
        self.apply_operation_btn.setObjectName("primaryButton")
        self.apply_operation_btn.setIcon(line_icon("check", "#FFFFFF", "#FFFFFF"))
        self.apply_operation_btn.setIconSize(QSize(18, 18))
        self.apply_operation_btn.setMinimumWidth(120)
        self.delete_selected_btn = QPushButton("删除选中")
        self.delete_selected_btn.setObjectName("dangerButton")
        self.delete_selected_btn.setIcon(line_icon("trash", "#FFFFFF", "#FFFFFF"))
        self.delete_selected_btn.setIconSize(QSize(18, 18))
        self.delete_selected_btn.setMinimumWidth(120)
        self.breadcrumb = QLabel("首页 > 文件管理")
        self.breadcrumb.setObjectName("breadcrumb")
        self.breadcrumb.setMinimumWidth(150)
        self.operation_combo.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        self.apply_operation_btn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        self.delete_selected_btn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)

        back.clicked.connect(self.go_back)
        refresh.clicked.connect(self.refresh_files)
        home.clicked.connect(self.go_home)
        new_folder.clicked.connect(self.create_new_folder)
        self.apply_operation_btn.clicked.connect(self.apply_selected_operation)
        self.delete_selected_btn.clicked.connect(self.delete_selected_items)
        layout.addWidget(back)
        layout.addWidget(refresh)
        layout.addWidget(home)
        layout.addWidget(new_folder)
        layout.addSpacing(12)
        layout.addWidget(self.breadcrumb)
        layout.addSpacing(12)
        layout.addWidget(self.operation_combo)
        layout.addWidget(self.apply_operation_btn)
        layout.addWidget(self.delete_selected_btn)
        layout.addStretch(1)
        return bar

    def _tool_button(self, pixmap: QStyle.StandardPixmap, tooltip: str) -> QToolButton:
        button = QToolButton()
        button.setObjectName("toolIconButton")
        icons = {
            QStyle.StandardPixmap.SP_ArrowBack: "back",
            QStyle.StandardPixmap.SP_BrowserReload: "refresh",
            QStyle.StandardPixmap.SP_DirHomeIcon: "home",
            QStyle.StandardPixmap.SP_FileDialogNewFolder: "folder_plus",
        }
        icon_name = icons.get(pixmap)
        button.setIcon(line_icon(icon_name, "#2C2C2E") if icon_name else self.style().standardIcon(pixmap))
        button.setIconSize(QSize(20, 20))
        button.setToolTip(tooltip)
        button.setFixedSize(38, 38)
        return button

    def _polish_table(self, table: QTableWidget, row_height: int = 50) -> None:
        table.setObjectName("dataTable")
        table.setShowGrid(False)
        table.setWordWrap(False)
        table.verticalHeader().setDefaultSectionSize(row_height)
        table.horizontalHeader().setFixedHeight(46)
        table.horizontalHeader().setHighlightSections(False)

    def _page_count(self) -> int:
        total = len(self.files)
        return max(1, (total + self.page_size - 1) // self.page_size) if total else 1

    def _visible_files(self) -> List[Dict[str, Any]]:
        start = self.file_page * self.page_size
        end = start + self.page_size
        return self.files[start:end]

    def _file_item_at_row(self, row: int) -> Optional[Dict[str, Any]]:
        if self.table_mode != "files":
            return None
        visible = self._visible_files()
        if 0 <= row < len(visible):
            return visible[row]
        return None

    def _sync_current_page_checks(self) -> None:
        if self.table_mode != "files" or not hasattr(self, "table"):
            return
        visible = self._visible_files()
        for row, item in enumerate(visible):
            fid = item.get("fid")
            check = self.table.item(row, 0)
            if not fid or not check:
                continue
            if check.checkState() == Qt.CheckState.Checked:
                self.checked_file_ids.add(fid)
            else:
                self.checked_file_ids.discard(fid)

    def _current_page_rows(self) -> List[tuple]:
        return [(row, item) for row, item in enumerate(self._visible_files())]

    def _update_page_toggle_button(self) -> None:
        if not hasattr(self, "select_page_folders_btn"):
            return
        if self.table_mode != "files":
            self.select_page_folders_btn.setText("全选")
            self.select_page_folders_btn.setEnabled(False)
            return

        page_rows = self._current_page_rows()
        if not page_rows:
            self.select_page_folders_btn.setText("全选")
            self.select_page_folders_btn.setEnabled(False)
            return

        all_checked = True
        for row, _item in page_rows:
            check = self.table.item(row, 0)
            if not check or check.checkState() != Qt.CheckState.Checked:
                all_checked = False
                break
        self.select_page_folders_btn.setEnabled(True)
        self.select_page_folders_btn.setText("取消全选" if all_checked else "全选")

    def _update_file_footer(self) -> None:
        if not hasattr(self, "file_footer"):
            return
        visible = self.content_stack.currentWidget() is self.file_list_page and self.table_mode == "files"
        self.file_footer.setVisible(visible)
        if not visible:
            return
        total = len(self.files)
        page_count = self._page_count()
        if self.file_page >= page_count:
            self.file_page = max(0, page_count - 1)
        start = self.file_page * self.page_size + 1 if total else 0
        end = min((self.file_page + 1) * self.page_size, total)
        self.page_summary_label.setText("%d-%d / %d" % (start, end, total) if total else "0 / 0")
        self.page_info_label.setText("第 %d / %d 页" % (self.file_page + 1, page_count))
        self.page_prev_btn.setEnabled(self.file_page > 0)
        self.page_next_btn.setEnabled(self.file_page + 1 < page_count)
        self._update_page_toggle_button()

    def _handle_table_item_changed(self, item: QTableWidgetItem) -> None:
        if self.table_mode != "files":
            return
        if item.column() != 0:
            return
        visible = self._visible_files()
        row = item.row()
        if not (0 <= row < len(visible)):
            return
        fid = visible[row].get("fid")
        if fid:
            if item.checkState() == Qt.CheckState.Checked:
                self.checked_file_ids.add(fid)
            else:
                self.checked_file_ids.discard(fid)
        self._update_page_toggle_button()

    def goto_prev_page(self) -> None:
        if self.file_page <= 0:
            return
        self._sync_current_page_checks()
        self.file_page -= 1
        self._render_files_page()

    def goto_next_page(self) -> None:
        if self.file_page + 1 >= self._page_count():
            return
        self._sync_current_page_checks()
        self.file_page += 1
        self._render_files_page()

    def toggle_select_current_page(self) -> None:
        if self.table_mode != "files":
            return
        page_rows = self._current_page_rows()
        if not page_rows:
            return
        all_checked = all(
            self.table.item(row, 0) and self.table.item(row, 0).checkState() == Qt.CheckState.Checked
            for row, _item in page_rows
        )
        affected = 0
        for row, item in page_rows:
            check = self.table.item(row, 0)
            fid = item.get("fid")
            if check:
                check.setCheckState(Qt.CheckState.Unchecked if all_checked else Qt.CheckState.Checked)
            if fid:
                if all_checked:
                    self.checked_file_ids.discard(fid)
                else:
                    self.checked_file_ids.add(fid)
            affected += 1
        if affected:
            if all_checked:
                self.append_log("已取消全选当前页的 %d 个项目" % affected)
            else:
                self.append_log("已全选当前页的 %d 个项目" % affected)
        self._update_file_footer()

    def _render_files_page(self) -> None:
        self.table_mode = "files"
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["", "文件名", "FID", "大小", "修改时间", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 48)
        self.table.setColumnWidth(2, 300)
        self.table.setColumnWidth(5, 180)

        visible = self._visible_files()
        self.table.setRowCount(len(visible))
        for row, item in enumerate(visible):
            fid = item.get("fid") or ""
            check = QTableWidgetItem()
            check.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            check.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            check.setCheckState(Qt.CheckState.Checked if fid and fid in self.checked_file_ids else Qt.CheckState.Unchecked)
            self.table.setItem(row, 0, check)

            name = QTableWidgetItem(item.get("file_name") or "未命名")
            name.setIcon(line_icon("folder_solid" if self._is_folder(item) else "file", "#8E8E93"))
            self.table.setItem(row, 1, name)
            self.table.setItem(row, 2, QTableWidgetItem(fid))
            self.table.setItem(row, 3, QTableWidgetItem(self._format_size(item.get("size"))))
            self.table.setItem(row, 4, QTableWidgetItem(self._format_time(item.get("updated_at") or item.get("created_at"))))
            self.table.setCellWidget(row, 5, self._row_actions(row))
            self.table.setRowHeight(row, 44)
        self.hint.setText("当前目录共 %d 项" % len(self.files))
        self._update_file_footer()

    def _reload_accounts(self) -> None:
        self.accounts = self.store.list_accounts()
        self.account_combo.blockSignals(True)
        self.account_combo.clear()
        if not self.accounts:
            self.account_combo.addItem(line_icon("account", "#8E8E93"), "未登录", "")
        else:
            active_index = 0
            for index, account in enumerate(self.accounts):
                self.account_combo.addItem(line_icon("account", "#3A3A3C"), account.name, account.account_id)
                if account.active:
                    active_index = index
            self.account_combo.setCurrentIndex(active_index)
        self.account_combo.blockSignals(False)
        self._load_current_client()
        if hasattr(self, "account_table"):
            self.render_settings_page()

    def _account_changed(self) -> None:
        account_id = self.account_combo.currentData()
        if account_id:
            self.store.set_active(account_id)
        self._load_current_client()
        self.refresh_files()

    def _load_current_client(self) -> None:
        account_id = self.account_combo.currentData()
        if not account_id:
            self.client = None
            return
        cookie = self.store.get_cookie(account_id)
        self.client = QuarkClient(cookie or "", self.append_log) if cookie else None

    def open_login(self) -> None:
        dialog = LoginDialog(self.store, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self._reload_accounts()
            if hasattr(self, "settings_page") and self.content_stack.currentWidget() is self.settings_page:
                self.show_settings()
            else:
                self.refresh_files()

    def create_new_folder(self) -> None:
        if not self._require_client():
            return
        name, ok = QInputDialog.getText(self, "新建文件夹", "请输入文件夹名称：")
        if not ok or not name.strip():
            return
        name = name.strip()

        def on_done(data: Any) -> None:
            self.append_log("文件夹 \"%s\" 创建成功" % name)
            self.refresh_files()

        self._start_api_call(self.client.create_folder, name, self.current_folder, on_success=on_done)

    def refresh_files(self) -> None:
        if not self._require_client():
            return
        self._set_active_nav("files")
        self.content_stack.setCurrentWidget(self.file_list_page)
        self.table_mode = "files"
        self._update_file_footer()
        self.breadcrumb.setText("首页 > 文件管理 > %s" % self.current_folder)
        self._start_api_call(self.client.list_files, self.current_folder, on_success=self.populate_files)  # type: ignore

    def search_files(self) -> None:
        keyword = self.search_input.text().strip()
        if not keyword:
            self.refresh_files()
            return
        if not self._require_client():
            return
        self._set_active_nav("files")
        self.content_stack.setCurrentWidget(self.file_list_page)
        self.table_mode = "files"
        self._update_file_footer()
        self.breadcrumb.setText("首页 > 搜索 > %s" % keyword)
        self._start_api_call(self.client.search_files, keyword, "root", on_success=self.populate_files)  # type: ignore

    def load_my_shares(self) -> None:
        if not self._require_client():
            return
        self._set_active_nav("myshare")
        self.content_stack.setCurrentWidget(self.file_list_page)
        self.table_mode = "shares"
        self._update_file_footer()
        self.breadcrumb.setText("首页 > 我的分享")
        self._start_api_call(self.client.get_share_list, on_success=self.populate_shares)  # type: ignore

    def populate_files(self, files: List[Dict[str, Any]]) -> None:
        self.files = files or []
        self.file_page = 0
        self.checked_file_ids = set()
        self._render_files_page()

    def populate_shares(self, shares: List[Dict[str, Any]]) -> None:
        self.table_mode = "shares"
        self.files = []
        self.checked_file_ids = set()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["标题", "链接", "创建时间", "统计"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setRowCount(len(shares or []))
        for row, item in enumerate(shares or []):
            self.table.setItem(row, 0, QTableWidgetItem(item.get("title") or "未命名分享"))
            self.table.setItem(row, 1, QTableWidgetItem(item.get("share_url") or ""))
            self.table.setItem(row, 2, QTableWidgetItem(self._format_time(item.get("create_time"))))
            stats = "访问 %s / 保存 %s / 下载 %s" % (
                item.get("visit_count", 0),
                item.get("save_count", 0),
                item.get("download_count", 0),
            )
            self.table.setItem(row, 3, QTableWidgetItem(stats))
            self.table.setRowHeight(row, 50)
        self.hint.setText("我的分享共 %d 项" % len(shares or []))
        self._update_file_footer()

    def _row_actions(self, row: int) -> QWidget:
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(6, 4, 6, 4)
        layout.setSpacing(6)
        share_btn = QPushButton("分享")
        rename_btn = QPushButton("重命名")
        share_btn.setIcon(line_icon("share", "#007AFF", "#007AFF"))
        rename_btn.setIcon(line_icon("rename", "#007AFF", "#007AFF"))
        for button in [share_btn, rename_btn]:
            button.setIconSize(QSize(16, 16))
            button.setFixedHeight(26)
            button.setObjectName("rowActionButton")
            layout.addWidget(button)
        layout.addStretch(1)
        share_btn.clicked.connect(lambda: self.share_row(row))
        rename_btn.clicked.connect(lambda: self.rename_row(row))
        return widget

    def _check_row(self, row: int) -> None:
        item = self.table.item(row, 0)
        if item and row < len(self.files):
            item.setCheckState(Qt.CheckState.Checked)
            self.append_log("已加入待操作：%s" % self.files[row].get("file_name"))

    def selected_files(self) -> List[Dict[str, Any]]:
        self._sync_current_page_checks()
        if self.checked_file_ids:
            return [item for item in self.files if item.get("fid") in self.checked_file_ids]
        rows = sorted({index.row() for index in self.table.selectionModel().selectedRows()})
        visible = self._visible_files()
        return [visible[row] for row in rows if row < len(visible)]

    def apply_selected_operation(self) -> None:
        action = self.operation_combo.currentData()
        files = self.selected_files()
        if not action:
            QMessageBox.information(self, "请选择操作类型", "先在下拉框里选择要加入的操作类型。")
            return
        if not files:
            QMessageBox.information(self, "未选择", "先勾选或选中要操作的文件。")
            return
        if action == "share_queue":
            added = 0
            for item in files:
                if self.add_share_item(item, show=False):
                    added += 1
            self.show_share_workspace()
            self.append_log("已加入 %d 个批量分享文件" % added)
            return
        if action == "copy_source":
            added = 0
            for item in files:
                if self.add_copy_source(item, show=False):
                    added += 1
            self.show_copy_workspace()
            self.append_log("已加入 %d 个被复制项目" % added)
            return
        if action == "copy_target":
            added = 0
            skipped = 0
            for item in files:
                if self.add_copy_target(item, show=len(files) == 1):
                    added += 1
                else:
                    skipped += 1
            if skipped and added == 0:
                QMessageBox.information(self, "没有加入目标", "“复制到”只能加入文件夹。要复制的文件请加入到“被复制”。")
            self.show_copy_workspace()
            self.append_log("已加入 %d 个复制目标文件夹" % added)
            return
        if action == "rename_queue":
            added = 0
            for item in files:
                if self.add_rename_item(item, show=False):
                    added += 1
            self.show_rename_workspace()
            self.append_log("已加入 %d 个批量重命名项目" % added)
            return
        if action == "move_queue":
            added = 0
            for item in files:
                if self.add_move_item(item, show=False):
                    added += 1
            self.show_move_workspace()
            self.append_log("已加入 %d 个批量移动项目" % added)
            return
        if action == "move_target":
            self.set_move_target(files[0])
            self.show_move_workspace()
            return
        if action == "delete_items":
            self.delete_items(files)

    def delete_selected_items(self) -> None:
        files = self.selected_files()
        if not files:
            QMessageBox.information(self, "未选择", "先勾选或选中要删除的文件或文件夹。")
            return
        self.delete_items(files)

    def delete_items(self, items: List[Dict[str, Any]]) -> None:
        deletable = [item for item in items if item.get("fid")]
        skipped = len(items) - len(deletable)
        if not deletable:
            QMessageBox.information(self, "没有可删除项目", "当前选中项里没有带 FID 的可删除文件或文件夹。")
            return

        preview = "\n".join("• %s" % (item.get("file_name") or item.get("fid")) for item in deletable[:8])
        if len(deletable) > 8:
            preview += "\n..."
        message = "确定要删除 %d 个项目吗？\n\n%s" % (len(deletable), preview)
        if skipped:
            message += "\n\n已自动跳过 %d 个缺少 FID 的项目。" % skipped
        confirm = QMessageBox.question(
            self,
            "确认删除",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            return

        jobs = [
            BatchJob(
                "delete",
                "删除 %s" % (item.get("file_name") or item.get("fid")),
                {"fid": item.get("fid"), "title": item.get("file_name") or ""},
            )
            for item in deletable
        ]
        self.start_jobs(jobs, 5, 10)

    def set_copy_source_from_row(self, row: int) -> None:
        item = self._file_item_at_row(row)
        if item:
            self.add_copy_source(item)
            self.show_copy_workspace()

    def add_copy_target_from_row(self, row: int) -> None:
        item = self._file_item_at_row(row)
        if item and self.add_copy_target(item):
            self.show_copy_workspace()

    def add_copy_source(self, item: Dict[str, Any], show: bool = True) -> bool:
        if not item.get("fid"):
            if show:
                QMessageBox.warning(self, "无法加入", "这个文件缺少 FID，不能加入被复制。")
            return False
        fid = item.get("fid")
        if any(source.get("fid") == fid for source in self.copy_sources):
            if show:
                QMessageBox.information(self, "已存在", "这个项目已经在被复制列表里。")
            return False
        self.copy_sources.append(item)
        self.append_log("已加入被复制：%s" % (item.get("file_name") or item.get("fid")))
        self.render_copy_workspace()
        return True

    def add_copy_target(self, item: Dict[str, Any], show: bool = True) -> bool:
        if not self._is_folder(item):
            if show:
                QMessageBox.information(self, "只能选择文件夹", "“复制到”目标必须是文件夹。")
            return False
        fid = item.get("fid")
        if not fid:
            if show:
                QMessageBox.warning(self, "无法加入", "这个文件夹缺少 FID，不能作为复制目标。")
            return False
        if any(target.get("fid") == fid for target in self.copy_targets):
            if show:
                QMessageBox.information(self, "已存在", "这个目标文件夹已经在复制到列表里。")
            return False
        self.copy_targets.append(item)
        self.copy_target_status[fid] = "等待"
        self.copy_target_progress.pop(fid, None)
        self.append_log("已加入复制到：%s" % (item.get("file_name") or fid))
        self.render_copy_workspace()
        return True

    def show_copy_workspace(self) -> None:
        self._set_active_nav("copy")
        self.content_stack.setCurrentWidget(self.copy_page)
        self.breadcrumb.setText("首页 > 批量复制")
        self.render_copy_workspace()

    def render_copy_workspace(self) -> None:
        if not hasattr(self, "copy_source_table"):
            return
        self.copy_summary.setText(
            "被复制 %d 项  ·  目标文件夹 %d 个"
            % (len(self.copy_sources), len(self.copy_targets))
        )
        self.copy_source_table.setRowCount(len(self.copy_sources))
        for row, source in enumerate(self.copy_sources):
            fid = source.get("fid") or ""
            name = QTableWidgetItem(source.get("file_name") or fid or "未命名")
            name.setIcon(line_icon("folder_solid" if self._is_folder(source) else "file", "#8E8E93"))
            self.copy_source_table.setItem(row, 0, name)
            self.copy_source_table.setCellWidget(
                row,
                1,
                self._mini_button("移除", lambda checked=False, source_id=fid: self.remove_copy_source(source_id), "dangerButton"),
            )
            self.copy_source_table.setRowHeight(row, 44)

        self.copy_target_table.setRowCount(len(self.copy_targets))
        for row, target in enumerate(self.copy_targets):
            fid = target.get("fid") or ""
            name = QTableWidgetItem(target.get("file_name") or fid)
            name.setIcon(line_icon("folder_solid", "#8E8E93"))
            self.copy_target_table.setItem(row, 0, name)
            status_text = self.copy_target_status.get(fid, "等待")
            status = self._status_item(
                status_text,
                "success"
                if status_text == "成功"
                else "danger"
                if status_text == "失败" or "部分失败" in status_text
                else "warning"
                if "执行" in status_text
                else "muted",
            )
            self.copy_target_table.setItem(row, 1, status)
            self.copy_target_table.setCellWidget(
                row,
                2,
                self._mini_button("移除", lambda checked=False, folder_id=fid: self.remove_copy_target(folder_id), "dangerButton"),
            )
            self.copy_target_table.setRowHeight(row, 44)

    def _mini_button(self, text: str, slot: Callable[..., Any], style_name: str) -> QWidget:
        box = QWidget()
        layout = QVBoxLayout(box)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(0)
        button = QPushButton(text)
        button.setObjectName(style_name)
        button.setFixedHeight(24)
        button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        button.setStyleSheet("margin-top: -1px;")
        button.clicked.connect(slot)
        row.addStretch(1)
        row.addWidget(button)
        row.addStretch(1)
        layout.addLayout(row)
        layout.addStretch(1)
        return box

    def _centered_icon_text(self, text: str, icon: QIcon) -> QWidget:
        box = QWidget()
        layout = QHBoxLayout(box)
        layout.setContentsMargins(4, 3, 4, 3)
        layout.setSpacing(4)
        layout.addStretch(1)
        icon_label = QLabel()
        icon_label.setPixmap(icon.pixmap(QSize(18, 18)))
        text_label = QLabel(text)
        text_label.setToolTip(text)
        layout.addWidget(icon_label)
        layout.addWidget(text_label)
        layout.addStretch(1)
        return box

    def _action_buttons(
        self,
        buttons: List[tuple],
        centered: bool = False,
        button_height: int = 24,
        margins: tuple = (4, 3, 4, 3),
        spacing: int = 4,
    ) -> QWidget:
        box = QWidget()
        box.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout = QHBoxLayout(box)
        layout.setContentsMargins(*margins)
        layout.setSpacing(spacing)
        if centered:
            layout.addStretch(1)
        for text, slot, style_name, enabled in buttons:
            button = QPushButton(text)
            button.setObjectName(style_name)
            button.setFixedHeight(button_height)
            button.setStyleSheet(f"min-height: {button_height}px; max-height: {button_height}px;")
            button.setMinimumWidth(max(52, 18 + len(text) * 14))
            button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
            button.setEnabled(enabled)
            button.clicked.connect(slot)
            layout.addWidget(button)
        layout.addStretch(1)
        return box

    def remove_copy_source(self, source_id: str) -> None:
        self.copy_sources = [item for item in self.copy_sources if item.get("fid") != source_id]
        self.render_copy_workspace()

    def remove_copy_target(self, folder_id: str) -> None:
        self.copy_targets = [item for item in self.copy_targets if item.get("fid") != folder_id]
        self.copy_target_status.pop(folder_id, None)
        self.copy_target_progress.pop(folder_id, None)
        self.render_copy_workspace()

    def clear_copy_workspace(self) -> None:
        self.copy_sources = []
        self.copy_targets = []
        self.copy_target_status = {}
        self.copy_target_progress = {}
        self.render_copy_workspace()

    def run_copy_workspace(self) -> None:
        if not self.copy_sources:
            QMessageBox.information(self, "缺少被复制文件", "先在文件管理里选中文件或文件夹，并加入到被复制。")
            return
        if not self.copy_targets:
            QMessageBox.information(self, "缺少复制目标", "先在文件管理里选择文件夹，并加入到复制到。")
            return
        jobs = []
        source_items = [item for item in self.copy_sources if item.get("fid")]
        self.copy_target_progress = {}
        for target in self.copy_targets:
            target_fid = target.get("fid")
            if not target_fid:
                continue
            valid_sources = [source for source in source_items if source.get("fid") and source.get("fid") != target_fid]
            total = len(valid_sources)
            if total <= 0:
                self.copy_target_status[target_fid] = "无可复制源"
                continue
            self.copy_target_progress[target_fid] = {"total": total, "done": 0, "failed": 0}
            self.copy_target_status[target_fid] = "等待执行 0/%d" % total
            for source in valid_sources:
                source_fid = source.get("fid")
                source_name = source.get("file_name") or source_fid
                jobs.append(
                    BatchJob(
                        "copy",
                        "复制 %s 到 %s" % (source_name, target.get("file_name") or target_fid),
                        {
                            "fid": source_fid,
                            "source_name": source_name,
                            "target_folder_id": target_fid,
                        },
                    )
                )
        self.render_copy_workspace()
        if not jobs:
            QMessageBox.information(self, "没有可执行任务", "当前源和目标没有形成可执行的复制组合。")
            return
        low = self.copy_delay_min
        high = max(self.copy_delay_max.value(), low)
        self.start_jobs(jobs, low, high)

    def add_rename_item(self, item: Dict[str, Any], show: bool = True) -> bool:
        fid = item.get("fid")
        if not fid:
            if show:
                QMessageBox.warning(self, "无法加入", "这个项目缺少 FID，不能加入批量重命名。")
            return False
        if any(existing.get("fid") == fid for existing in self.rename_items):
            if show:
                QMessageBox.information(self, "已存在", "这个项目已经在批量重命名列表里。")
            return False
        queued = dict(item)
        queued["new_name"] = item.get("file_name") or ""
        self.rename_items.append(queued)
        self.rename_status[fid] = "等待"
        self.render_rename_workspace()
        return True

    def show_rename_workspace(self) -> None:
        self._set_active_nav("rename")
        self.content_stack.setCurrentWidget(self.rename_page)
        self.breadcrumb.setText("首页 > 批量重命名")
        self.render_rename_workspace()

    def render_rename_workspace(self) -> None:
        if not hasattr(self, "rename_table"):
            return
        renamed = sum(1 for value in self.rename_status.values() if value == "成功")
        self.rename_summary.setText("已加入 %d 项，已完成 %d 项" % (len(self.rename_items), renamed))
        self.rename_table.setRowCount(len(self.rename_items))
        readonly_flags = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled
        editable_flags = readonly_flags | Qt.ItemFlag.ItemIsEditable
        for row, item in enumerate(self.rename_items):
            fid = item.get("fid") or ""
            old_name = QTableWidgetItem(item.get("file_name") or fid)
            old_name.setIcon(line_icon("folder_solid" if self._is_folder(item) else "file", "#8E8E93"))
            old_name.setFlags(readonly_flags)
            self.rename_table.setItem(row, 0, old_name)
            new_name = QTableWidgetItem("   " + (item.get("new_name") or item.get("file_name") or ""))
            new_name.setFlags(editable_flags)
            self.rename_table.setItem(row, 1, new_name)
            status_text = self.rename_status.get(fid, "等待")
            status = self._status_item(
                status_text,
                "success" if status_text == "成功" else "danger" if status_text in ("失败", "名称为空") else "warning" if status_text in ("等待执行", "未修改") else "muted",
            )
            status.setFlags(readonly_flags)
            self.rename_table.setItem(row, 2, status)
            self.rename_table.setCellWidget(
                row,
                3,
                self._mini_button("移除", lambda checked=False, file_id=fid: self.remove_rename_item(file_id), "dangerButton"),
            )
            self.rename_table.setRowHeight(row, 44)

    def _sync_rename_table_edits(self) -> None:
        if not hasattr(self, "rename_table"):
            return
        for row in range(min(self.rename_table.rowCount(), len(self.rename_items))):
            self.rename_items[row]["new_name"] = self._table_text(self.rename_table.item(row, 1))

    def remove_rename_item(self, fid: str) -> None:
        self.rename_items = [item for item in self.rename_items if item.get("fid") != fid]
        self.rename_status.pop(fid, None)
        self.render_rename_workspace()

    def clear_rename_workspace(self) -> None:
        self.rename_items = []
        self.rename_status = {}
        self.render_rename_workspace()

    def run_rename_workspace(self) -> None:
        if not self.rename_items:
            QMessageBox.information(self, "没有重命名任务", "先在文件管理里勾选文件，再加入到批量重命名。")
            return
        self._sync_rename_table_edits()
        jobs = []
        for item in self.rename_items:
            fid = item.get("fid")
            old_name = item.get("file_name") or ""
            new_name = (item.get("new_name") or "").strip()
            if not fid:
                continue
            if not new_name:
                self.rename_status[fid] = "名称为空"
                continue
            if new_name == old_name:
                self.rename_status[fid] = "未修改"
                continue
            self.rename_status[fid] = "等待执行"
            jobs.append(
                BatchJob(
                    "rename",
                    "重命名 %s -> %s" % (old_name or fid, new_name),
                    {"fid": fid, "new_name": new_name},
                )
            )
        self.render_rename_workspace()
        if not jobs:
            QMessageBox.information(self, "没有可执行任务", "没有需要重命名的项目；请在“新文件名”列里修改名称。")
            return
        low = self.rename_delay_min
        high = max(self.rename_delay_max.value(), low)
        self.start_jobs(jobs, low, high)

    def add_move_item(self, item: Dict[str, Any], show: bool = True) -> bool:
        fid = item.get("fid")
        if not fid:
            if show:
                QMessageBox.warning(self, "无法加入", "这个项目缺少 FID，不能加入批量移动。")
            return False
        if any(existing.get("fid") == fid for existing in self.move_items):
            if show:
                QMessageBox.information(self, "已存在", "这个项目已经在批量移动列表里。")
            return False
        self.move_items.append(dict(item))
        self.move_status[fid] = "等待"
        self.render_move_workspace()
        return True

    def set_move_target(self, item: Dict[str, Any]) -> bool:
        if not self._is_folder(item):
            QMessageBox.information(self, "只能选择文件夹", "移动目标必须是文件夹。")
            return False
        fid = item.get("fid")
        if not fid:
            QMessageBox.warning(self, "无法设置", "这个文件夹缺少 FID，不能作为移动目标。")
            return False
        self.move_target = item
        if hasattr(self, "move_target_fid"):
            self.move_target_fid.setText(fid)
        self.append_log("已设为移动目标：%s" % (item.get("file_name") or fid))
        self.render_move_workspace()
        return True

    def show_move_workspace(self) -> None:
        self._set_active_nav("move")
        self.content_stack.setCurrentWidget(self.move_page)
        self.breadcrumb.setText("首页 > 批量移动")
        self.render_move_workspace()

    def render_move_workspace(self) -> None:
        if not hasattr(self, "move_table"):
            return
        target_fid = self.move_target_fid.text().strip() if hasattr(self, "move_target_fid") else ""
        target_name = ""
        if self.move_target and self.move_target.get("fid") == target_fid:
            target_name = self.move_target.get("file_name") or ""
        if target_name:
            self.move_target_name.setText(target_name)
        else:
            self.move_target_name.setText("已填写目标" if target_fid else "未设置目标")
        target_display = target_name or target_fid or "-"
        moved = sum(1 for value in self.move_status.values() if value == "成功")
        self.move_summary.setText("已加入 %d 项，目标位置：%s，已完成 %d 项" % (len(self.move_items), target_display, moved))

        self.move_table.setRowCount(len(self.move_items))
        for row, item in enumerate(self.move_items):
            fid = item.get("fid") or ""
            name = QTableWidgetItem("   " + (item.get("file_name") or fid))
            name.setIcon(line_icon("folder_solid" if self._is_folder(item) else "file", "#8E8E93"))
            self.move_table.setItem(row, 0, name)
            target_item = QTableWidgetItem(target_display)
            target_item.setIcon(line_icon("folder_solid", "#8E8E93"))
            self.move_table.setItem(row, 1, target_item)
            status_text = self.move_status.get(fid, "等待")
            status = self._status_item(
                status_text,
                "success" if status_text == "成功" else "danger" if status_text == "失败" else "warning" if "执行" in status_text else "muted",
            )
            self.move_table.setItem(row, 2, status)
            self.move_table.setCellWidget(
                row,
                3,
                self._mini_button("移除", lambda checked=False, file_id=fid: self.remove_move_item(file_id), "dangerButton"),
            )
            self.move_table.setRowHeight(row, 44)

    def remove_move_item(self, fid: str) -> None:
        self.move_items = [item for item in self.move_items if item.get("fid") != fid]
        self.move_status.pop(fid, None)
        self.render_move_workspace()

    def clear_move_workspace(self) -> None:
        self.move_items = []
        self.move_status = {}
        self.move_target = None
        if hasattr(self, "move_target_fid"):
            self.move_target_fid.clear()
        self.render_move_workspace()

    def run_move_workspace(self) -> None:
        if not self.move_items:
            QMessageBox.information(self, "没有移动任务", "先在文件管理里勾选文件，再加入到批量移动。")
            return
        target_fid = self.move_target_fid.text().strip()
        if not target_fid:
            QMessageBox.information(self, "缺少目标位置", "请在文件管理里选择文件夹并设为移动到位置，或手动填写目标文件夹 FID。")
            return
        jobs = []
        for item in self.move_items:
            fid = item.get("fid")
            if not fid:
                continue
            if fid == target_fid:
                self.move_status[fid] = "跳过自身"
                continue
            self.move_status[fid] = "等待执行"
            jobs.append(
                BatchJob(
                    "move",
                    "移动 %s 到 %s" % (item.get("file_name") or fid, self.move_target_name.text() or target_fid),
                    {"fid": fid, "target_folder_id": target_fid},
                )
            )
        self.render_move_workspace()
        if not jobs:
            QMessageBox.information(self, "没有可执行任务", "没有可移动的项目。")
            return
        low = self.move_delay_min
        high = max(self.move_delay_max.value(), low)
        self.start_jobs(jobs, low, high)

    def add_share_item(self, item: Dict[str, Any], show: bool = True) -> bool:
        fid = item.get("fid")
        if not fid:
            if show:
                QMessageBox.warning(self, "无法加入", "这个项目缺少 FID，不能加入批量分享。")
            return False
        if any(existing.get("fid") == fid for existing in self.share_items):
            if show:
                QMessageBox.information(self, "已存在", "这个项目已经在批量分享列表里。")
            return False
        self.share_items.append(item)
        self.share_results[fid] = {"status": "等待", "share_url": "", "password": "", "error": ""}
        self.render_share_workspace()
        return True

    def show_share_workspace(self) -> None:
        self._set_active_nav("share")
        self.content_stack.setCurrentWidget(self.share_page)
        self.breadcrumb.setText("首页 > 批量分享")
        self.render_share_workspace()

    def render_share_workspace(self) -> None:
        if not hasattr(self, "share_table"):
            return
        shared = sum(1 for result in self.share_results.values() if result.get("status") == "成功")
        self.share_summary.setText("已加入 %d 项，已生成 %d 条分享链接" % (len(self.share_items), shared))
        self.share_table.setRowCount(len(self.share_items))
        for row, item in enumerate(self.share_items):
            fid = item.get("fid") or ""
            result = self.share_results.get(fid, {})
            name = QTableWidgetItem(item.get("file_name") or fid)
            name.setIcon(line_icon("folder_solid" if self._is_folder(item) else "file", "#8E8E93"))
            self.share_table.setItem(row, 0, name)
            status_text = result.get("status") or "等待"
            status = self._status_item(
                status_text,
                "success" if status_text == "成功" else "danger" if status_text == "失败" else "warning" if "执行" in status_text else "muted",
                result.get("error") or "",
            )
            self.share_table.setItem(row, 1, status)
            self.share_table.setItem(row, 2, QTableWidgetItem(result.get("share_url") or ""))
            self.share_table.setItem(row, 3, QTableWidgetItem(result.get("password") or ""))
            self.share_table.setCellWidget(
                row,
                4,
                self._mini_button("移除", lambda checked=False, file_id=fid: self.remove_share_item(file_id), "dangerButton"),
            )
            self.share_table.setRowHeight(row, 44)

    def remove_share_item(self, fid: str) -> None:
        self.share_items = [item for item in self.share_items if item.get("fid") != fid]
        self.share_results.pop(fid, None)
        self.render_share_workspace()

    def clear_share_workspace(self) -> None:
        self.share_items = []
        self.share_results = {}
        self.render_share_workspace()

    def run_share_workspace(self) -> None:
        if not self.share_items:
            QMessageBox.information(self, "没有分享任务", "先在文件管理里勾选文件，再加入到批量分享。")
            return
        jobs = []
        for item in self.share_items:
            fid = item.get("fid")
            if not fid:
                continue
            self.share_results[fid] = {
                "status": "等待执行",
                "share_url": self.share_results.get(fid, {}).get("share_url", ""),
                "password": self.share_results.get(fid, {}).get("password", ""),
                "error": "",
            }
            jobs.append(
                BatchJob(
                    "share",
                    "分享 %s" % (item.get("file_name") or fid),
                    {
                        "fid": fid,
                        "file_name": item.get("file_name") or fid,
                        "title": item.get("file_name") or "夸克分享",
                        "expire_days": 0,
                        "password": "",
                    },
                )
            )
        self.render_share_workspace()
        low = self.share_delay_min
        high = max(self.share_delay_max.value(), low)
        self.start_jobs(jobs, low, high)

    def export_share_results(self) -> None:
        if not self.share_items:
            QMessageBox.information(self, "没有数据", "批量分享列表为空。")
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception:
            QMessageBox.warning(self, "缺少依赖", "缺少 openpyxl，请先运行：python -m pip install -r requirements.txt")
            return

        default_name = "批量分享结果_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")
        path, _ = QFileDialog.getSaveFileName(self, "导出 Excel", default_name, "Excel 工作簿 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "批量分享结果"
        headers = ["文件名", "FID", "分享链接", "提取码", "状态", "错误信息"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for item in self.share_items:
            fid = item.get("fid") or ""
            result = self.share_results.get(fid, {})
            sheet.append(
                [
                    item.get("file_name") or fid,
                    fid,
                    result.get("share_url") or "",
                    result.get("password") or "",
                    result.get("status") or "",
                    result.get("error") or "",
                ]
            )
        widths = [34, 34, 60, 12, 12, 36]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(path)
        self.append_log("已导出批量分享结果：%s" % path)
        QMessageBox.information(self, "导出完成", "Excel 已导出。")

    def show_transfer_workspace(self) -> None:
        self._set_active_nav("transfer")
        self.content_stack.setCurrentWidget(self.transfer_page)
        self.breadcrumb.setText("首页 > 批量转存")
        self.render_transfer_workspace()

    def parse_transfer_lines(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for raw_line in self.transfer_links.toPlainText().splitlines():
            line = raw_line.strip()
            if not line:
                continue
            match = re.search(r"https?://pan\.quark\.cn/s/[a-zA-Z0-9]+", line)
            if not match:
                continue
            code_match = re.search(r"(?:提取码|密码|passcode|code)[:：\s]*([a-zA-Z0-9]+)", line)
            passcode = code_match.group(1) if code_match else ""
            rows.append(
                {
                    "game_name": "",
                    "file_name": "",
                    "share_url": match.group(0),
                    "passcode": passcode,
                    "target_folder_id": self.transfer_target_folder.text().strip() or "root",
                    "status": "等待",
                    "title": "",
                    "auto_share_url": "",
                    "error": "",
                }
            )
        return rows

    def run_transfer_workspace(self) -> None:
        rows = self.transfer_results if self.transfer_results else self.parse_transfer_lines()
        if not rows:
            QMessageBox.information(self, "没有任务", "没有识别到有效的夸克分享链接。")
            return
        self.transfer_results = rows
        self._sync_transfer_table_edits()
        jobs = []
        default_target_folder = self.transfer_target_folder.text().strip() or "root"
        kind = "transfer_share" if self.transfer_auto_share.isChecked() else "transfer"
        for index, row in enumerate(self.transfer_results):
            row["status"] = "等待执行"
            row["auto_share_url"] = row.get("auto_share_url", "")
            row["error"] = ""
            target_folder = row.get("target_folder_id") or default_target_folder
            row["target_folder_id"] = target_folder
            jobs.append(
                BatchJob(
                    kind,
                    "转存 %s" % (row.get("file_name") or row["share_url"]),
                    {
                        "row_index": index,
                        "share_url": row["share_url"],
                        "passcode": row["passcode"],
                        "target_folder_id": target_folder,
                        "expire_days": 0,
                        "share_delay_min": 4,
                        "share_delay_max": 6,
                        "share_retries": 3,
                    },
                )
            )
        self.render_transfer_workspace()
        low = self.transfer_delay_min
        high = max(self.transfer_delay_max.value(), low)
        if kind == "transfer_share":
            self.append_log("转存并分享节奏：转存前随机等待 %d-%d 秒，转存完成后每次分享前等待 4-6 秒，最多重试 3 次。" % (low, high))
        self.start_jobs(jobs, low, high)

    def render_transfer_workspace(self) -> None:
        if not hasattr(self, "transfer_table"):
            return
        success_count = sum(1 for item in self.transfer_results if (item.get("status") or "") == "成功")
        share_count = sum(1 for item in self.transfer_results if item.get("auto_share_url"))
        self.transfer_summary.setText(
            "当前 %d 条记录，成功 %d 条，已生成分享链接 %d 条"
            % (len(self.transfer_results), success_count, share_count)
        )
        self.transfer_table.setRowCount(len(self.transfer_results))
        for row, item in enumerate(self.transfer_results):
            self.transfer_table.setItem(row, 0, QTableWidgetItem(item.get("file_name") or ""))
            self.transfer_table.setItem(row, 1, QTableWidgetItem(item.get("share_url") or ""))
            status_text = item.get("status") or "等待"
            tone = "muted"
            if status_text == "成功":
                tone = "success"
            elif status_text in ("失败", "分享失败"):
                tone = "danger"
            elif "执行" in status_text:
                tone = "warning"
            status = self._status_item(status_text, tone, item.get("error") or "")
            self.transfer_table.setItem(row, 2, status)
            self.transfer_table.setItem(row, 3, QTableWidgetItem(item.get("auto_share_url") or ""))
            self.transfer_table.setItem(row, 4, QTableWidgetItem(item.get("error") or ""))
            self.transfer_table.setRowHeight(row, 44)

    def clear_transfer_results(self) -> None:
        self.transfer_results = []
        self.transfer_links.clear()
        self.render_transfer_workspace()

    def _sync_transfer_table_edits(self) -> None:
        if not hasattr(self, "transfer_table"):
            return
        for row in range(min(self.transfer_table.rowCount(), len(self.transfer_results))):
            item = self.transfer_results[row]
            item["file_name"] = self._table_text(self.transfer_table.item(row, 0))
            item["share_url"] = self._table_text(self.transfer_table.item(row, 1))

    def _table_text(self, item: Optional[QTableWidgetItem]) -> str:
        return item.text().strip() if item else ""

    def view_import_template(self) -> None:
        import os
        template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "导入Excel模板.xlsx")
        if not os.path.exists(template_path):
            QMessageBox.warning(self, "模板不存在", "未找到导入Excel模板文件：\n%s" % template_path)
            return
        try:
            os.startfile(template_path)
        except Exception as exc:
            QMessageBox.warning(self, "打开失败", "无法打开模板文件：%s" % str(exc))

    def import_transfer_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, "导入转存 Excel", "", "Excel 工作簿 (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
        except Exception:
            QMessageBox.warning(self, "缺少依赖", "缺少 openpyxl，请先运行：python -m pip install -r requirements.txt")
            return
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            header = [self._cell_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            column_map = self._map_transfer_excel_columns(header)
            rows: List[Dict[str, Any]] = []
            default_target = self.transfer_target_folder.text().strip() or "root"
            for values in sheet.iter_rows(min_row=2, values_only=True):
                url = self._cell_text(values[column_map["url"]]) if column_map["url"] < len(values) else ""
                if not url:
                    continue
                match = re.search(r"https?://pan\.quark\.cn/s/[a-zA-Z0-9]+", url)
                if not match:
                    continue
                passcode = ""
                if "passcode" in column_map and column_map["passcode"] < len(values):
                    passcode = self._cell_text(values[column_map["passcode"]])
                target_folder = default_target
                if "target_folder_id" in column_map and column_map["target_folder_id"] < len(values):
                    target_folder = self._cell_text(values[column_map["target_folder_id"]]) or default_target
                rows.append(
                    {
                        "game_name": self._cell_text(values[column_map["game_name"]]) if "game_name" in column_map and column_map["game_name"] < len(values) else "",
                        "file_name": self._cell_text(values[column_map["file_name"]]) if "file_name" in column_map and column_map["file_name"] < len(values) else "",
                        "share_url": match.group(0),
                        "passcode": passcode,
                        "target_folder_id": target_folder,
                        "status": "等待",
                        "title": "",
                        "auto_share_url": "",
                        "error": "",
                    }
                )
        except Exception as exc:
            QMessageBox.warning(self, "导入失败", str(exc))
            return

        self.transfer_results = rows
        self.transfer_links.setPlainText("\n".join(row["share_url"] for row in rows))
        self.transfer_auto_share.setChecked(True)
        self.render_transfer_workspace()
        self.append_log("已从 Excel 导入 %d 条转存并分享任务" % len(rows))
        if not rows:
            QMessageBox.information(self, "没有可导入数据", "没有找到有效的夸克链接数据行。")

    def export_transfer_results(self) -> None:
        if not self.transfer_results:
            QMessageBox.information(self, "没有数据", "当前没有可导出的转存结果。")
            return
        self._sync_transfer_table_edits()
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception:
            QMessageBox.warning(self, "缺少依赖", "缺少 openpyxl，请先运行：python -m pip install -r requirements.txt")
            return
        default_name = "转存分享结果_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")
        path, _ = QFileDialog.getSaveFileName(self, "导出转存分享结果", default_name, "Excel 工作簿 (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "转存分享结果"
        # headers = ["游戏名", "夸克链接", "文件名称", "状态"]
        headers = ["文件名称", "夸克链接", "状态"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in self.transfer_results:
            sheet.append(
                [
                    # row.get("game_name") or "",
                    row.get("file_name") or row.get("title") or "",
                    row.get("auto_share_url") or "",
                    row.get("status") or "",
                ]
            )
        # widths = [28, 62, 42, 14]
        widths = [42, 62, 14]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(path)
        self.append_log("已导出转存分享结果：%s" % path)
        QMessageBox.information(self, "导出完成", "Excel 已导出。")

    def _map_transfer_excel_columns(self, header: List[str]) -> Dict[str, int]:
        normalized = {self._normalize_header(name): index for index, name in enumerate(header)}
        mapping = {
            "url": self._find_header_index(normalized, ["夸克链接", "链接", "quark链接", "url", "shareurl"]),
        }
        game_name_index = self._find_header_index(normalized, ["游戏名", "游戏名称", "game", "gamename"], required=False)
        if game_name_index is not None:
            mapping["game_name"] = game_name_index
        file_name_index = self._find_header_index(normalized, ["文件名称", "文件名", "资源名称", "filename"], required=False)
        if file_name_index is not None:
            mapping["file_name"] = file_name_index
        passcode_index = self._find_header_index(normalized, ["提取码", "密码", "passcode", "code"], required=False)
        if passcode_index is not None:
            mapping["passcode"] = passcode_index
        target_index = self._find_header_index(
            normalized,
            ["转存地址FID", "保存到文件夹FID", "目标文件夹FID", "目标FID", "targetfolderid", "targetfid"],
            required=False,
        )
        if target_index is not None:
            mapping["target_folder_id"] = target_index
        return mapping

    def _find_header_index(self, normalized: Dict[str, int], candidates: List[str], required: bool = True) -> Optional[int]:
        for candidate in candidates:
            key = self._normalize_header(candidate)
            if key in normalized:
                return normalized[key]
        if required:
            raise ValueError("Excel 缺少列：%s" % " / ".join(candidates))
        return None

    def _normalize_header(self, value: str) -> str:
        text = str(value or "").strip().lower()
        for token in (" ", "\t", "\n", "\r", "_", "-", "/", "\\", "（", "）", "(", ")"):
            text = text.replace(token, "")
        return text

    def _cell_text(self, value: Any) -> str:
        return "" if value is None else str(value).strip()

    def batch_share_selected(self) -> None:
        files = self.selected_files()
        if not files:
            QMessageBox.information(self, "未选择", "先勾选要分享的文件或文件夹。")
            return
        added = 0
        for item in files:
            if self.add_share_item(item, show=False):
                added += 1
        self.show_share_workspace()
        self.append_log("已加入 %d 个批量分享文件" % added)

    def share_row(self, row: int) -> None:
        item = self._file_item_at_row(row)
        if not item:
            return
        job = BatchJob(
            "share",
            "分享 %s" % item.get("file_name"),
            {"fid": item.get("fid"), "title": item.get("file_name") or "夸克分享"},
        )
        self.start_jobs([job], 5, 10)

    def rename_selected(self) -> None:
        files = self.selected_files()
        if len(files) != 1:
            QMessageBox.information(self, "请选择一个项目", "批量重命名规则下一版做；当前先支持单个重命名。")
            return
        self.rename_item(files[0])

    def rename_row(self, row: int) -> None:
        item = self._file_item_at_row(row)
        if item:
            self.rename_item(item)

    def rename_item(self, item: Dict[str, Any]) -> None:
        new_name, ok = QInputDialog.getText(self, "重命名", "新名称：", text=item.get("file_name") or "")
        if not ok or not new_name.strip():
            return
        if not self._require_client():
            return
        self._start_api_call(
            self.client.rename_file,  # type: ignore
            item["fid"],
            new_name.strip(),
            on_success=lambda _: self.refresh_files(),
        )

    def open_transfer_dialog(self) -> None:
        self.show_transfer_workspace()

    def open_copy_dialog(self) -> None:
        selected = self.selected_files()
        for item in selected:
            self.add_copy_source(item, show=False)
        self.show_copy_workspace()

    def copy_row(self, row: int) -> None:
        item = self._file_item_at_row(row)
        if item:
            self.add_copy_source(item)
            self.show_copy_workspace()

    def move_selected_placeholder(self) -> None:
        selected = self.selected_files()
        if selected:
            for item in selected:
                self.add_move_item(item, show=False)
        self.show_move_workspace()

    def show_settings(self) -> None:
        self._set_active_nav("settings")
        self.content_stack.setCurrentWidget(self.settings_page)
        self.breadcrumb.setText("首页 > 账号池")
        self.render_settings_page()

    def _selected_table_rows(self, table: QTableWidget) -> List[int]:
        rows = set()
        if table.selectionModel():
            rows.update(index.row() for index in table.selectionModel().selectedRows())
        if not rows:
            rows.update(item.row() for item in table.selectedItems())
        return sorted(rows)

    def _set_active_nav(self, key: str) -> None:
        self.current_nav = key
        for nav_key, button in self.nav_buttons.items():
            active = nav_key == key
            if button.property("active") != active:
                button.setProperty("active", active)
                button.style().unpolish(button)
                button.style().polish(button)
                button.update()

    def _status_item(self, text: str, tone: str = "muted", tooltip: str = "") -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        colors = {
            "muted": "#8E8E93",
            "success": "#24A148",
            "danger": "#D92D20",
            "warning": "#C97A10",
            "info": "#0A84FF",
        }
        item.setForeground(QColor(colors.get(tone, colors["muted"])))
        if tooltip:
            item.setToolTip(tooltip)
        return item

    def show_trash_workspace(self) -> None:
        if not self._require_client():
            return
        self._set_active_nav("trash")
        self.content_stack.setCurrentWidget(self.trash_page)
        self.breadcrumb.setText("首页 > 回收站")
        self._start_api_call(self.client.list_recycle_files, on_success=self.populate_recycle_files)  # type: ignore

    def populate_recycle_files(self, items: List[Dict[str, Any]]) -> None:
        self.recycle_items = items or []
        valid_ids = {item.get("record_id") for item in self.recycle_items if item.get("record_id")}
        self.recycle_status = {key: value for key, value in self.recycle_status.items() if key in valid_ids}
        self.hint.setText("回收站共 %d 项，可恢复或彻底删除。" % len(self.recycle_items))
        self.render_trash_workspace()

    def render_trash_workspace(self) -> None:
        if not hasattr(self, "trash_table"):
            return
        self.recycle_visible_items = self._filtered_recycle_items()
        selected_count = len(self._selected_recycle_items())
        self.trash_summary.setText(
            "回收站共 %d 项，当前显示 %d 项，已选 %d 项"
            % (len(self.recycle_items), len(self.recycle_visible_items), selected_count)
        )
        self.trash_table.blockSignals(True)
        self.trash_table.clearSelection()
        self.trash_table.setRowCount(len(self.recycle_visible_items))
        for row, item in enumerate(self.recycle_visible_items):
            record_id = item.get("record_id") or ""
            fid = item.get("fid") or ""
            name = QTableWidgetItem(item.get("file_name") or fid or "未命名")
            name.setIcon(line_icon("folder_solid" if self._is_folder(item) else "file", "#8E8E93"))
            name.setToolTip(item.get("file_name") or fid or "未命名")
            self.trash_table.setItem(row, 0, name)
            deleted_at = self._format_time(item.get("move_recycle_at") or item.get("updated_at"))
            deleted_item = QTableWidgetItem(deleted_at)
            deleted_item.setToolTip(deleted_at)
            deleted_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.trash_table.setItem(row, 1, deleted_item)
            remain = QTableWidgetItem(self._format_recycle_remain(item.get("invisible_remain")))
            remain.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.trash_table.setItem(row, 2, remain)
            status_text = self.recycle_status.get(record_id, "待处理")
            if status_text in ("已恢复",):
                status = self._status_item(status_text, "success")
            elif status_text in ("已删除", "失败"):
                status = self._status_item(status_text, "danger")
            elif status_text in ("恢复中", "删除中"):
                status = self._status_item(status_text, "warning")
            else:
                status = self._status_item(status_text, "muted")
            self.trash_table.setItem(row, 3, status)
            self.trash_table.setCellWidget(
                row,
                4,
                self._action_buttons(
                    [
                        ("恢复", lambda checked=False, rec_id=record_id: self.recover_recycle_row(rec_id), "successButton", True),
                        ("删除", lambda checked=False, rec_id=record_id: self.remove_recycle_row(rec_id), "dangerButton", True),
                    ],
                    centered=True,
                    button_height=24,
                    margins=(4, 10, 4, 10),
                ),
            )
            self.trash_table.setRowHeight(row, 44)
        self.trash_table.blockSignals(False)
        self._update_trash_selection_state()

    def _filtered_recycle_items(self) -> List[Dict[str, Any]]:
        if not hasattr(self, "trash_filter_combo"):
            return list(self.recycle_items)
        mode = self.trash_filter_combo.currentData()
        visible: List[Dict[str, Any]] = []
        for item in self.recycle_items:
            remain_seconds = self._recycle_remain_seconds(item.get("invisible_remain"))
            days = remain_seconds / 86400.0 if remain_seconds > 0 else 0.0
            if mode == "lt1" and not (remain_seconds > 0 and days <= 1):
                continue
            if mode == "lt3" and not (remain_seconds > 0 and days <= 3):
                continue
            if mode == "lt7" and not (remain_seconds > 0 and days <= 7):
                continue
            if mode == "gt7" and not (days > 7):
                continue
            if mode == "expired" and not (remain_seconds <= 0):
                continue
            visible.append(item)
        return visible

    def _recycle_remain_seconds(self, value: Any) -> float:
        try:
            return float(value or 0) / 1000.0
        except Exception:
            return 0.0

    def _format_recycle_remain(self, value: Any) -> str:
        seconds = self._recycle_remain_seconds(value)
        if seconds <= 0:
            return "已过期"
        days = seconds / 86400.0
        if days >= 1:
            return "%.1f天" % days
        hours = seconds / 3600.0
        if hours >= 1:
            return "%.1f小时" % hours
        return "%.0f分钟" % max(seconds / 60.0, 1.0)

    def _selected_recycle_items(self) -> List[Dict[str, Any]]:
        rows = self._selected_table_rows(self.trash_table) if hasattr(self, "trash_table") else []
        return [self.recycle_visible_items[row] for row in rows if 0 <= row < len(self.recycle_visible_items)]

    def _update_trash_selection_state(self) -> None:
        if not hasattr(self, "trash_table"):
            return
        total = len(self.recycle_visible_items)
        selected = len(self._selected_table_rows(self.trash_table))
        if hasattr(self, "trash_select_btn"):
            self.trash_select_btn.setText("取消全选" if total and selected >= total else "全选")
        if hasattr(self, "trash_summary"):
            self.trash_summary.setText(
                "回收站共 %d 项，当前显示 %d 项，已选 %d 项"
                % (len(self.recycle_items), total, selected)
            )

    def toggle_select_recycle_visible(self) -> None:
        if not hasattr(self, "trash_table"):
            return
        total = len(self.recycle_visible_items)
        selected = len(self._selected_table_rows(self.trash_table))
        if total and selected >= total:
            self.trash_table.clearSelection()
        else:
            self.trash_table.selectAll()
        self._update_trash_selection_state()

    def recover_selected_recycle(self) -> None:
        items = self._selected_recycle_items()
        if not items:
            QMessageBox.information(self, "未选择", "先在回收站里选中要恢复的文件或文件夹。")
            return
        self._start_recycle_jobs(items, "recycle_recover")

    def remove_selected_recycle(self) -> None:
        items = self._selected_recycle_items()
        if not items:
            QMessageBox.information(self, "未选择", "先在回收站里选中要彻底删除的文件或文件夹。")
            return
        names = "、".join((item.get("file_name") or "未命名") for item in items[:3])
        if len(items) > 3:
            names += " 等 %d 项" % len(items)
        if QMessageBox.question(self, "确认彻底删除", "将永久删除：%s\n这个操作无法撤销，是否继续？" % names) != QMessageBox.StandardButton.Yes:
            return
        self._start_recycle_jobs(items, "recycle_remove")

    def recover_recycle_row(self, record_id: str) -> None:
        item = next((entry for entry in self.recycle_items if entry.get("record_id") == record_id), None)
        if item:
            self._start_recycle_jobs([item], "recycle_recover")

    def remove_recycle_row(self, record_id: str) -> None:
        item = next((entry for entry in self.recycle_items if entry.get("record_id") == record_id), None)
        if not item:
            return
        name = item.get("file_name") or "未命名"
        if QMessageBox.question(self, "确认彻底删除", "将永久删除 %s，是否继续？" % name) != QMessageBox.StandardButton.Yes:
            return
        self._start_recycle_jobs([item], "recycle_remove")

    def clear_recycle_bin(self) -> None:
        if not self.recycle_items:
            QMessageBox.information(self, "回收站为空", "当前没有可清理的回收站内容。")
            return
        if QMessageBox.question(
            self,
            "确认清空回收站",
            "将彻底删除回收站中的 %d 项内容。\n这个操作无法撤销，是否继续？" % len(self.recycle_items),
        ) != QMessageBox.StandardButton.Yes:
            return
        self._start_recycle_jobs(list(self.recycle_items), "recycle_remove")

    def _start_recycle_jobs(self, items: List[Dict[str, Any]], kind: str) -> None:
        jobs = []
        pending_status = "恢复中" if kind == "recycle_recover" else "删除中"
        for item in items:
            record_id = item.get("record_id")
            if not record_id:
                continue
            self.recycle_status[record_id] = pending_status
            name = item.get("file_name") or record_id
            jobs.append(
                BatchJob(
                    kind,
                    ("恢复 %s" if kind == "recycle_recover" else "彻底删除 %s") % name,
                    {"record_id": record_id, "file_name": name},
                )
            )
        if not jobs:
            QMessageBox.warning(self, "无法执行", "选中的回收站记录缺少 record_id。")
            return
        self.render_trash_workspace()
        low = self.trash_delay_min
        high = max(self.trash_delay_max.value(), low)
        self.start_jobs(jobs, low, high)

    def render_settings_page(self) -> None:
        if not hasattr(self, "account_table"):
            return
        active_name = next((account.name for account in self.accounts if account.active), "未登录")
        self.account_pool_summary.setText("已保存 %d 个账号，当前账号：%s" % (len(self.accounts), active_name))
        self.account_table.setRowCount(len(self.accounts))
        for row, account in enumerate(self.accounts):
            self.account_table.setCellWidget(
                row,
                0,
                self._centered_icon_text(account.name, line_icon("account", "#3A3A3C", "#5AC8FA")),
            )
            current = self._status_item("当前" if account.active else "-", "info" if account.active else "muted")
            self.account_table.setItem(row, 1, current)
            health = self.account_health.get(account.account_id, {})
            status_text = health.get("status") or ("待检测" if self.store.get_cookie(account.account_id) else "缺少登录态")
            if status_text == "可用":
                status_item = self._status_item(status_text, "success", health.get("detail") or status_text)
            elif status_text in ("异常", "缺少登录态"):
                status_item = self._status_item(status_text, "danger", health.get("detail") or status_text)
            else:
                status_item = self._status_item(status_text, "warning", health.get("detail") or status_text)
            self.account_table.setItem(row, 2, status_item)
            self.account_table.setItem(row, 3, QTableWidgetItem(self._format_time(account.created_at)))
            cookie_state = self._status_item(
                "已保存" if self.store.get_cookie(account.account_id) else "缺失",
                "success" if self.store.get_cookie(account.account_id) else "danger",
            )
            self.account_table.setItem(row, 4, cookie_state)
            self.account_table.setCellWidget(
                row,
                5,
                self._action_buttons(
                    [
                        ("当前" if account.active else "切换", lambda checked=False, account_id=account.account_id: self.set_active_account_from_pool(account_id), "smallGrayButton" if account.active else "primaryButton", not account.active),
                        ("改名", lambda checked=False, account_id=account.account_id: self.rename_account_from_pool(account_id), "smallGrayButton", True),
                        ("检查", lambda checked=False, account_id=account.account_id: self.check_single_account(account_id), "successButton", True),
                        ("删除", lambda checked=False, account_id=account.account_id: self.delete_account_from_pool(account_id), "dangerButton", True),
                    ],
                    centered=True,
                    button_height=34,
                    margins=(4, 0, 4, 4),
                    spacing=6,
                ),
            )
            self.account_table.setRowHeight(row, 44)

    def set_active_account_from_pool(self, account_id: str) -> None:
        self.store.set_active(account_id)
        self._reload_accounts()
        account = next((item for item in self.accounts if item.account_id == account_id), None)
        self.append_log("已切换当前账号：%s" % (account.name if account else account_id))
        self.render_settings_page()

    def rename_account_from_pool(self, account_id: str) -> None:
        account = next((item for item in self.accounts if item.account_id == account_id), None)
        if not account:
            return
        new_name, ok = QInputDialog.getText(self, "重命名账号", "账号名称：", text=account.name)
        if not ok:
            return
        try:
            self.store.rename_account(account_id, new_name)
        except ValueError as exc:
            QMessageBox.warning(self, "无法重命名", str(exc))
            return
        self._reload_accounts()
        self.append_log("已重命名账号：%s" % new_name.strip())
        self.render_settings_page()

    def delete_account_from_pool(self, account_id: str) -> None:
        account = next((item for item in self.accounts if item.account_id == account_id), None)
        if not account:
            return
        if QMessageBox.question(self, "删除账号", "确认删除账号 %s 吗？已保存的登录态也会一起移除。" % account.name) != QMessageBox.StandardButton.Yes:
            return
        self.store.delete_account(account_id)
        self.account_health.pop(account_id, None)
        self._reload_accounts()
        self.append_log("已删除账号：%s" % account.name)
        self.render_settings_page()

    def check_single_account(self, account_id: str) -> None:
        self._start_api_call(self._probe_accounts, [account_id], on_success=self._apply_account_health)

    def refresh_account_pool_statuses(self) -> None:
        if not self.accounts:
            QMessageBox.information(self, "没有账号", "先添加至少一个账号。")
            return
        self._start_api_call(self._probe_accounts, on_success=self._apply_account_health)

    def _probe_accounts(self, account_ids: Optional[List[str]] = None) -> List[Dict[str, str]]:
        wanted = set(account_ids or [])
        results: List[Dict[str, str]] = []
        for account in self.store.list_accounts():
            if wanted and account.account_id not in wanted:
                continue
            cookie = self.store.get_cookie(account.account_id)
            if not cookie:
                results.append({"account_id": account.account_id, "status": "缺少登录态", "detail": "没有找到可用 Cookie"})
                continue
            try:
                client = QuarkClient(cookie)
                client.list_files("root", 1)
                results.append({"account_id": account.account_id, "status": "可用", "detail": "根目录读取正常"})
            except Exception as exc:
                results.append({"account_id": account.account_id, "status": "异常", "detail": str(exc)[:160]})
        return results

    def _apply_account_health(self, results: List[Dict[str, str]]) -> None:
        checked = 0
        for item in results or []:
            account_id = item.get("account_id") or ""
            if not account_id:
                continue
            self.account_health[account_id] = {"status": item.get("status") or "未知", "detail": item.get("detail") or ""}
            checked += 1
        self.render_settings_page()
        if checked:
            self.append_log("账号检测完成：%d 个账号" % checked)

    def focus_log(self) -> None:
        self._set_active_nav("log")
        self.log_box.setFocus()

    def open_row(self, row: int, column: int) -> None:
        item = self._file_item_at_row(row)
        if not item:
            return
        if not self._is_folder(item):
            return
        fid = item.get("fid")
        if not fid:
            return
        self.folder_history.append(self.current_folder)
        self.current_folder = fid
        self.refresh_files()

    def go_back(self) -> None:
        if self.content_stack.currentWidget() in (
            self.copy_page,
            self.share_page,
            self.transfer_page,
            self.rename_page,
            self.move_page,
            self.trash_page,
            self.settings_page,
        ):
            self.content_stack.setCurrentWidget(self.file_list_page)
            self.breadcrumb.setText("首页 > 文件管理 > %s" % self.current_folder)
            self._set_active_nav("files")
            return
        if self.folder_history:
            self.current_folder = self.folder_history.pop()
            self.refresh_files()
            return
        self.go_home()

    def go_home(self) -> None:
        self.folder_history = []
        self.current_folder = "root"
        self.refresh_files()

    def start_jobs(self, jobs: List[BatchJob], delay_min: int, delay_max: int) -> None:
        if not jobs:
            return
        if not self._require_client():
            return
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "任务运行中", "已有批量任务正在运行，请先停止或等待完成。")
            return
        self.progress.setRange(0, len(jobs))
        self.progress.setValue(0)
        self.stop_btn.setEnabled(True)
        self.last_job_kinds = {job.kind for job in jobs}
        self.append_log("开始批量任务：%d 个，随机延时 %d-%d 秒" % (len(jobs), delay_min, delay_max))
        self.worker = TaskWorker(self.client, jobs, delay_min, delay_max)  # type: ignore
        self.worker.log.connect(self.append_log)
        self.worker.progress.connect(lambda done, total: self.progress.setValue(done))
        self.worker.failed.connect(lambda message: self.append_log("失败：" + message))
        self.worker.result.connect(self._handle_job_result)
        self.worker.finished_all.connect(self._jobs_finished)
        self.worker.start()

    def stop_worker(self) -> None:
        if self.worker:
            self.worker.cancel()
            self.append_log("正在停止任务...")

    def _handle_job_result(self, result: Dict[str, Any]) -> None:
        if result.get("kind") == "copy":
            payload = result.get("payload") or {}
            target_id = payload.get("target_folder_id")
            if target_id:
                progress = self.copy_target_progress.get(target_id)
                if progress:
                    progress["done"] = int(progress.get("done", 0)) + 1
                    if not result.get("ok"):
                        progress["failed"] = int(progress.get("failed", 0)) + 1
                    done = int(progress.get("done", 0))
                    total = int(progress.get("total", 0))
                    failed = int(progress.get("failed", 0))
                    if done >= total:
                        if failed <= 0:
                            self.copy_target_status[target_id] = "成功"
                        elif failed >= total:
                            self.copy_target_status[target_id] = "失败"
                        else:
                            self.copy_target_status[target_id] = "部分失败 %d/%d" % (failed, total)
                    else:
                        if failed > 0:
                            self.copy_target_status[target_id] = "执行中 %d/%d，失败 %d" % (done, total, failed)
                        else:
                            self.copy_target_status[target_id] = "执行中 %d/%d" % (done, total)
                else:
                    self.copy_target_status[target_id] = "成功" if result.get("ok") else "失败"
                self.render_copy_workspace()
        if result.get("kind") == "rename":
            payload = result.get("payload") or {}
            fid = payload.get("fid")
            if fid:
                self.rename_status[fid] = "成功" if result.get("ok") else "失败"
                self.render_rename_workspace()
        if result.get("kind") == "move":
            payload = result.get("payload") or {}
            fid = payload.get("fid")
            if fid:
                self.move_status[fid] = "成功" if result.get("ok") else "失败"
                self.render_move_workspace()
        if result.get("kind") in ("recycle_recover", "recycle_remove"):
            payload = result.get("payload") or {}
            record_id = payload.get("record_id")
            if record_id:
                if result.get("ok"):
                    self.recycle_status[record_id] = "已恢复" if result.get("kind") == "recycle_recover" else "已删除"
                else:
                    self.recycle_status[record_id] = "失败"
                self.render_trash_workspace()
        if result.get("kind") == "share":
            payload = result.get("payload") or {}
            fid = payload.get("fid")
            if fid:
                current = self.share_results.get(fid, {})
                if result.get("ok"):
                    data = result.get("data") or {}
                    current.update(
                        {
                            "status": "成功",
                            "share_url": self._find_nested_value(data, ["share_url", "url", "link"]) or "",
                            "password": self._find_nested_value(data, ["password", "passcode", "pwd", "share_pwd"]) or "",
                            "error": "",
                        }
                    )
                else:
                    current.update({"status": "失败", "error": result.get("error") or ""})
                self.share_results[fid] = current
                self.render_share_workspace()
        if result.get("kind") in ("transfer", "transfer_share"):
            payload = result.get("payload") or {}
            row_index = payload.get("row_index")
            if isinstance(row_index, int) and 0 <= row_index < len(self.transfer_results):
                row = self.transfer_results[row_index]
                if result.get("ok"):
                    data = result.get("data") or {}
                    row["title"] = data.get("title") or row.get("title") or ""
                    link = self._find_nested_value(data, ["share_url", "url", "link"]) or ""
                    row["auto_share_url"] = link
                    if result.get("kind") == "transfer_share":
                        if link:
                            row["status"] = "成功"
                            row["error"] = ""
                        else:
                            row["status"] = "分享失败"
                            row["error"] = data.get("share_error") or self._find_nested_value(data, ["message", "msg"]) or "转存成功，但没有生成分享链接"
                            self.append_log("分享失败：%s - %s" % (result.get("label"), row["error"]))
                    else:
                        row["status"] = "成功"
                        row["error"] = ""
                else:
                    row["status"] = "失败"
                    row["error"] = result.get("error") or ""
                self.render_transfer_workspace()
        if result.get("ok"):
            data = result.get("data") or {}
            link = data.get("share_url") or (data.get("share") or {}).get("share_url")
            if link:
                self.append_log("分享链接：%s" % link)
        else:
            self.append_log("任务失败：%s - %s" % (result.get("label"), result.get("error")))

    def _jobs_finished(self) -> None:
        kinds = set(self.last_job_kinds)
        self.last_job_kinds = set()
        self.stop_btn.setEnabled(False)
        self.append_log("批量任务结束")
        if self.content_stack.currentWidget() is self.copy_page:
            self.render_copy_workspace()
        elif self.content_stack.currentWidget() is self.share_page:
            self.render_share_workspace()
        elif self.content_stack.currentWidget() is self.transfer_page:
            self.render_transfer_workspace()
        elif self.content_stack.currentWidget() is self.rename_page:
            self.render_rename_workspace()
        elif self.content_stack.currentWidget() is self.move_page:
            self.render_move_workspace()
        elif self.content_stack.currentWidget() is self.trash_page or kinds.intersection({"recycle_recover", "recycle_remove"}):
            if self.client:
                self._start_api_call(self.client.list_recycle_files, on_success=self.populate_recycle_files)
        else:
            self.refresh_files()

    def _start_api_call(self, func: Callable[..., Any], *args: Any, on_success: Callable[[Any], None]) -> None:
        if self.api_thread and self.api_thread.isRunning():
            self.append_log("已有请求正在执行，请稍候。")
            return
        self.append_log("请求中...")
        self.api_thread = ApiThread(func, *args)
        self.api_thread.success.connect(on_success)
        self.api_thread.success.connect(lambda _: self.append_log("请求完成"))
        self.api_thread.error.connect(lambda message: QMessageBox.warning(self, "接口错误", message))
        self.api_thread.error.connect(lambda message: self.append_log("接口错误：" + message))
        self.api_thread.start()

    def _require_client(self) -> bool:
        if self.client:
            return True
        QMessageBox.information(self, "需要登录", "请先点击“添加账号”，扫码登录并保存登录态。")
        return False

    def append_log(self, message: str) -> None:
        now = time.strftime("%H:%M:%S")
        self.log_box.append("[%s] %s" % (now, message))

    def _find_nested_value(self, data: Any, keys: List[str]) -> str:
        if isinstance(data, dict):
            for key in keys:
                value = data.get(key)
                if value:
                    return str(value)
            for value in data.values():
                found = self._find_nested_value(value, keys)
                if found:
                    return found
        elif isinstance(data, list):
            for item in data:
                found = self._find_nested_value(item, keys)
                if found:
                    return found
        return ""

    def _format_size(self, size: Any) -> str:
        try:
            value = float(size or 0)
        except Exception:
            return "-"
        if value <= 0:
            return "-"
        units = ["B", "KB", "MB", "GB", "TB"]
        unit = 0
        while value >= 1024 and unit < len(units) - 1:
            value /= 1024
            unit += 1
        return "%.2f %s" % (value, units[unit])

    def _format_time(self, ts: Any) -> str:
        try:
            value = float(ts or 0)
        except Exception:
            return "-"
        if value <= 0:
            return "-"
        if value > 100000000000:
            value = value / 1000
        return time.strftime("%Y-%m-%d %H:%M", time.localtime(value))

    def _is_folder(self, item: Dict[str, Any]) -> bool:
        if "file" in item:
            value = item.get("file")
            if isinstance(value, bool):
                return not value
            if isinstance(value, (int, float)) and value in (0, 1):
                return not bool(value)
            if isinstance(value, str) and value.lower() in ("true", "false", "1", "0"):
                return value.lower() in ("false", "0")
        for key in ("is_dir", "dir", "is_folder", "folder"):
            value = item.get(key)
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)) and value in (0, 1):
                return bool(value)
            if isinstance(value, str) and value.lower() in ("true", "false", "1", "0"):
                return value.lower() in ("true", "1")

        file_type = item.get("file_type")
        if isinstance(file_type, str):
            lowered = file_type.lower()
            if lowered in ("folder", "dir", "directory"):
                return True
            if lowered in ("file", "video", "audio", "image", "doc", "archive"):
                return False
            if lowered.isdigit():
                return int(lowered) == 0
        if isinstance(file_type, (int, float)):
            return int(file_type) == 0
        return False

    def _apply_style(self) -> None:
        ui_families = preferred_ui_families()
        font = QFont(ui_families[0], 10)
        if hasattr(font, "setFamilies"):
            font.setFamilies(ui_families)
        font.setHintingPreference(QFont.HintingPreference.PreferFullHinting)
        self.setFont(font)
        font_stack = ", ".join('"%s"' % family for family in ui_families)
        self.setStyleSheet(
            """
            QMainWindow, QWidget#appRoot {
                background: #f5f5f7;
            }
            QWidget {
                color: #1d1d1f;
                font-size: 14px;
                font-family: %s;
            }
            #sidebar {
                background: #f7f7f9;
                border-right: none;
            }
            #sidebarHeader {
                background: #ffffff;
                border-bottom: 1px solid #e5e5ea;
            }
            #sidebarNav {
                background: #f7f7f9;
            }
            #brand {
                color: #111114;
                font-size: 23px;
                font-weight: 700;
                padding: 0;
            }
            #quota, #version, #mutedLabel {
                color: #86868b;
            }
            #sectionCaption {
                color: #6e6e73;
                font-size: 13px;
                padding: 0 0 2px 0;
            }
            #version {
                font-weight: 520;
                color: #8e8e93;
                padding: 10px 4px 2px 4px;
            }
            #pageTitle {
                color: #111114;
                font-size: 20px;
                font-weight: 640;
                padding: 2px 0 4px 0;
            }
            QPushButton {
                min-height: 34px;
                border-radius: 9px;
                padding: 0 14px;
                border: 1px solid #d9d9e2;
                background: #ffffff;
                color: #1d1d1f;
                font-weight: 540;
            }
            QPushButton:hover {
                background: #f8f8fa;
                border-color: #cfcfd7;
            }
            QPushButton:pressed {
                background: #efeff3;
            }
            QPushButton:disabled {
                background: #f1f1f3;
                border-color: #e2e2e7;
                color: #a1a1a8;
            }
            #navButton {
                text-align: left;
                min-height: 42px;
                border-radius: 10px;
                border: 1px solid transparent;
                background: transparent;
                padding: 0 12px;
                color: #2c2c2e;
                font-family: "Microsoft YaHei UI";
                font-size: 16px;
                font-weight: 500;
            }
            #navButton:hover {
                background: #ffffff;
                border-color: #ededf1;
            }
            #navButton[active="true"] {
                background: #ffffff;
                border-color: #dbe9ff;
                color: #0a84ff;
            }
            #primaryButton {
                background: #007aff;
                border-color: #007aff;
                color: #ffffff;
            }
            #primaryButton:hover {
                background: #006fe6;
                border-color: #006fe6;
            }
            #successButton {
                background: #34c759;
                border-color: #34c759;
                color: #ffffff;
            }
            #successButton:hover {
                background: #2eb94f;
                border-color: #2eb94f;
            }
            #dangerButton {
                background: #ff3b30;
                border-color: #ff3b30;
                color: #ffffff;
            }
            #dangerButton:hover {
                background: #ed3026;
                border-color: #ed3026;
            }
            #rowActionButton {
                min-height: 28px;
                border-radius: 7px;
                padding: 0 10px;
                background: #f4f8ff;
                color: #006edb;
                border: 1px solid #dfebff;
                font-size: 13px;
                font-weight: 560;
            }
            #rowActionButton:hover {
                background: #ebf3ff;
                border-color: #cfe3ff;
            }
            #rowActionButton:disabled {
                background: #f2f2f4;
                color: #b0b0b8;
                border-color: #e5e5ea;
            }
            #smallGrayButton {
                background: #f6f6f8;
                color: #3a3a3c;
                border: 1px solid #e4e4ea;
                padding: 0 10px;
            }
            #smallGrayButton:hover {
                background: #efeff3;
                border-color: #d7d7df;
            }
            #smallGrayButton:pressed {
                background: #e8e8ed;
            }
            #topbar, #toolbar {
                background: #ffffff;
                border-bottom: 1px solid #e5e5ea;
            }
            #fileFooter {
                background: #ffffff;
                border-top: 1px solid #e9e9ee;
            }
            #toolbar {
                background: #fbfbfd;
            }
            #breadcrumb {
                color: #3a3a3c;
                font-weight: 600;
                padding: 0 4px;
            }
            #statusPanel {
                background: #fbfbfd;
                border-top: 1px solid #e5e5ea;
            }
            QStackedWidget {
                background: #ffffff;
                border: none;
            }
            #logBox {
                background: #fcfcfd;
                border: 1px solid #e8e8ee;
                border-radius: 10px;
                color: #3a3a3c;
            }
            QLineEdit, QComboBox, QPlainTextEdit, QTextEdit, QSpinBox {
                background: #fcfcfd;
                border: 1px solid #dfdfe6;
                border-radius: 10px;
                padding: 7px 10px;
                selection-background-color: #007aff;
                selection-color: #ffffff;
            }
            QLineEdit:focus, QComboBox:focus, QPlainTextEdit:focus, QTextEdit:focus, QSpinBox:focus {
                border-color: #007aff;
            }
            QComboBox#filterCombo {
                min-height: 34px;
                padding: 0 10px;
            }
            QSpinBox#compactSpin {
                min-width: 34px;
                max-width: 34px;
                min-height: 34px;
                max-height: 34px;
                padding: 0 2px;
            }
            QComboBox {
                min-height: 34px;
            }
            QComboBox#topCombo {
                min-height: 34px;
                max-height: 34px;
                padding: 0 10px;
            }
            QComboBox::drop-down {
                border: none;
                width: 28px;
            }
            QSpinBox::up-button, QSpinBox::down-button {
                border: none;
                width: 18px;
                background: transparent;
            }
            QToolButton#toolIconButton {
                background: #ffffff;
                border: 1px solid #d8d8de;
                border-radius: 8px;
                padding: 0;
                color: #3a3a3c;
                font-size: 23px;
                font-weight: 560;
            }
            QToolButton#toolIconButton:hover {
                background: #f5f5f7;
                border-color: #c7c7cc;
            }
            QToolButton#toolIconButton:pressed {
                background: #ececf0;
            }
            QTableWidget {
                background: #fcfcfd;
                alternate-background-color: #ffffff;
                border: none;
                color: #1d1d1f;
                selection-background-color: #edf5ff;
                selection-color: #1d1d1f;
            }
            QTableWidget::item {
                border-bottom: 1px solid #f1f1f4;
                padding: 0 8px;
            }
            QTableWidget::item:selected {
                background: #edf5ff;
                color: #1d1d1f;
            }
            QHeaderView::section {
                background: #f8f8fb;
                border: none;
                border-bottom: 1px solid #e5e5ea;
                padding: 0 10px;
                color: #7a7a81;
                font-weight: 620;
            }
            QProgressBar {
                min-height: 12px;
                max-height: 12px;
                border: 1px solid #d8d8de;
                border-radius: 6px;
                text-align: center;
                background: #f2f2f4;
                color: transparent;
            }
            QProgressBar::chunk {
                background: #34c759;
                border-radius: 5px;
            }
            QSplitter::handle {
                background: #f5f5f7;
            }
            QSplitter::handle:vertical {
                height: 1px;
            }
            QSplitter::handle:horizontal {
                width: 1px;
            }
            QScrollBar:vertical {
                width: 10px;
                background: transparent;
                margin: 2px;
            }
            QScrollBar::handle:vertical {
                background: #c7c7cc;
                border-radius: 5px;
                min-height: 36px;
            }
            QScrollBar::handle:vertical:hover {
                background: #aeaeb2;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical,
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                height: 0;
                background: transparent;
            }
            QCheckBox {
                spacing: 8px;
                color: #3a3a3c;
                font-weight: 500;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border-radius: 4px;
                border: 1px solid #c7c7cc;
                background: #ffffff;
            }
            QCheckBox::indicator:checked {
                background: #007aff;
                border-color: #007aff;
            }
            """ % font_stack
        )
